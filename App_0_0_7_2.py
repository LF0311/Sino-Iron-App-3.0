# -*- coding: utf-8 -*-

import streamlit as st
import pandas as pd
import datetime
from PIL import Image, ImageDraw, ImageFont
import plotly.graph_objects as go
from plotly.subplots import make_subplots
import random
import json
import yaml
from yaml.loader import SafeLoader
import streamlit_authenticator as stauth
from streamlit_image_zoom import image_zoom
import psycopg2
import numpy as np
import os
import glob
from sqlalchemy import create_engine, text
from openpyxl import load_workbook
import openpyxl

import bcrypt

def generate_hashed_password(password: str) -> str:
    """生成 bcrypt 加密后的密码"""
    # 将密码编码为 bytes
    password_bytes = password.encode('utf-8')
    
    # 生成 salt 并加密
    salt = bcrypt.gensalt()
    hashed = bcrypt.hashpw(password_bytes, salt)
    
    # 返回字符串形式
    return hashed.decode('utf-8')

# ── PostgreSQL 连接配置 ──────────────────────────────────────────────────────────
DB_CONNECTION = 'postgresql://postgres:postgres@localhost:5432/postgres'


@st.cache_data(ttl=300)
def load_cvr_data_pg(cvr_name, start_date, end_date):
    """从 cvr_tracking 表查询指定 CVR 的数据"""
    try:
        engine = create_engine(DB_CONNECTION)
        query = text("""
            SELECT * FROM cvr_tracking
            WHERE cvr_name = :name
              AND time >= :start AND time <= :end
            ORDER BY time
        """)
        with engine.connect() as conn:
            df = pd.read_sql(query, conn, params={
                'name': cvr_name,
                'start': pd.Timestamp(start_date),
                'end': pd.Timestamp(end_date) + pd.Timedelta(days=1),
            })
        if not df.empty:
            df = df.rename(columns={'time': 'Current Timestamp'})
            df['Current Timestamp'] = pd.to_datetime(df['Current Timestamp'])
        return df
    except Exception as e:
        st.error(f"Error loading CVR data: {e}")
        return pd.DataFrame()


@st.cache_data(ttl=300)
def load_tripper_data_pg(tripper_name, start_date, end_date):
    """从 tripper_tracking 表查询指定布料机的数据"""
    try:
        engine = create_engine(DB_CONNECTION)
        query = text("""
            SELECT * FROM tripper_tracking
            WHERE tripper_name = :name
              AND time >= :start AND time <= :end
            ORDER BY time
        """)
        with engine.connect() as conn:
            df = pd.read_sql(query, conn, params={
                'name': tripper_name,
                'start': pd.Timestamp(start_date),
                'end': pd.Timestamp(end_date) + pd.Timedelta(days=1),
            })
        if not df.empty:
            df = df.rename(columns={'time': 'Current Timestamp'})
            df['Current Timestamp'] = pd.to_datetime(df['Current Timestamp'])
        return df
    except Exception as e:
        st.error(f"Error loading Tripper data: {e}")
        return pd.DataFrame()


@st.cache_data(ttl=300)
def load_silo_tracking_pg(silo_num, start_date, end_date):
    """从 silo_tracking 表查询指定料仓的数据"""
    try:
        engine = create_engine(DB_CONNECTION)
        query = text("""
            SELECT * FROM silo_tracking
            WHERE silo_num = :num
              AND time >= :start AND time <= :end
            ORDER BY time
        """)
        with engine.connect() as conn:
            df = pd.read_sql(query, conn, params={
                'num': int(silo_num),
                'start': pd.Timestamp(start_date),
                'end': pd.Timestamp(end_date) + pd.Timedelta(days=1),
            })
        if not df.empty:
            df = df.rename(columns={'time': 'Current Timestamp'})
            df['Current Timestamp'] = pd.to_datetime(df['Current Timestamp'])
        return df
    except Exception as e:
        st.error(f"Error loading silo tracking data: {e}")
        return pd.DataFrame()


@st.cache_data(ttl=300)
def load_mill_feed_pg(mill_num, start_date, end_date):
    """从 mill_feed 表查询指定磨机的数据"""
    try:
        engine = create_engine(DB_CONNECTION)
        query = text("""
            SELECT * FROM mill_feed
            WHERE mill_num = :num
              AND time >= :start AND time <= :end
            ORDER BY time
        """)
        with engine.connect() as conn:
            df = pd.read_sql(query, conn, params={
                'num': int(mill_num),
                'start': pd.Timestamp(start_date),
                'end': pd.Timestamp(end_date) + pd.Timedelta(days=1),
            })
        if not df.empty:
            df = df.rename(columns={'time': 'Current Timestamp'})
            df['Current Timestamp'] = pd.to_datetime(df['Current Timestamp'])
        return df
    except Exception as e:
        st.error(f"Error loading mill feed data: {e}")
        return pd.DataFrame()


def local_pvModel(file_name):
    st.markdown(
            f'<iframe src=' + file_name + ' height = "860" width = "100%"></iframe>',
            unsafe_allow_html=True,
    )


def generate_stockpile_data(dates):
    # Set a fixed seed to ensure consistent results across runs
    random.seed(42)

    # Generate Stockpile names
    stockpile_names = ["S1", "S2", "S3", "S4", "S5", "S6"]

    # Define reasonable ranges for each mineral property
    magfe_range = (18, 24)
    dtr_range = (29, 38)
    magfe_dtr_range = (20, 26)
    fe_dtr_range = (64, 75)
    fe_head_range = (28, 35)
    fe_tail_range = (8, 20)
    sio2_dtr_range = (1.99, 8.69)
    cao_head_range = (1.4, 1.89)
    mt_index_range = (75.1, 80.21)
    oxidation_value = 1  # Oxidation is a fixed value
    strata_values = ["J1", "J2", "J3", "J4", "J5", "J6"]
    pen_rate_range = (13, 28)
    p80_imt_range = (55, 71)

    # Initialize the data list
    stockpile_data = []

    # Generate data for each Stockpile
    for name in stockpile_names:
        data = {
            "Stockpile": name,
            "Date": [],
            "MagFe%": [],
            "DTR%": [],
            "MagFe_DTR%": [],
            "Fe_DTR%": [],
            "Fe_Head%": [],
            "Fe_Tail%": [],
            "SiO2_DTR%": [],
            "CaO_Head%": [],
            "MT_Index": [],
            "Oxidation": [],
            "Strata": [],
            "PEN_RATE": [],
            "P80_IMT_um": []
        }
        for date in dates:
            data["Date"].append(date.date())
            data["MagFe%"].append(random.uniform(*magfe_range))
            data["DTR%"].append(random.uniform(*dtr_range))
            data["MagFe_DTR%"].append(random.uniform(*magfe_dtr_range))
            data["Fe_DTR%"].append(random.uniform(*fe_dtr_range))
            data["Fe_Head%"].append(random.uniform(*fe_head_range))
            data["Fe_Tail%"].append(random.uniform(*fe_tail_range))
            data["SiO2_DTR%"].append(random.uniform(*sio2_dtr_range))
            data["CaO_Head%"].append(random.uniform(*cao_head_range))
            data["MT_Index"].append(random.uniform(*mt_index_range))
            data["Oxidation"].append(oxidation_value)
            data["Strata"].append(random.choice(strata_values))
            data["PEN_RATE"].append(random.uniform(*pen_rate_range))
            data["P80_IMT_um"].append(random.uniform(*p80_imt_range))
        stockpile_data.append(data)

    # Convert data to DataFrame
    stockpile_dfs = [pd.DataFrame(data) for data in stockpile_data]

    return stockpile_dfs


# Function to annotate the image
def annotate_image(image_path, t1, t2, t3, t4, t5, t6, t7, t8, t9, t10):
    # Load the image
    image = Image.open(image_path)
    draw = ImageDraw.Draw(image)

    # Define font and color
    try:
        font = ImageFont.truetype("arial.ttf", 40)  # Ensure the font is accessible or use default
    except IOError:
        font = ImageFont.load_default()

    color = (255, 0, 0)  # Red color for the text

    # Draw text on the image at specified positions
    draw.text((1080, 300), f"T1A={t1}s", fill=color, font=font)  # Replace with exact coordinates for t1
    draw.text((1260, 660), f"T2A={t2}s", fill=color, font=font)  # Replace with exact coordinates for t2
    draw.text((1000, 810), f"T3A={t3}s", fill=color, font=font)  # Replace with exact coordinates for t3
    draw.text((1000, 1000), f"T4A={t4}s", fill=color, font=font)  # Replace with exact coordinates for t4
    draw.text((1200, 1200), f"T5A={t5}s", fill=color, font=font)  # Replace with exact coordinates for t5
    draw.text((1450, 1050), f"T6A={t6}s", fill=color, font=font)  # Replace with exact coordinates for t6

    draw.text((800, 1750), f"T1B={t1}s", fill=color, font=font)  # Replace with exact coordinates for t1
    draw.text((880, 2070), f"T2B={t2}s", fill=color, font=font)  # Replace with exact coordinates for t2
    draw.text((740, 2250), f"T3B={t3}s", fill=color, font=font)  # Replace with exact coordinates for t3
    draw.text((760, 2480), f"T4B={t4}s", fill=color, font=font)  # Replace with exact coordinates for t4
    draw.text((1120, 2730), f"T5B={t5}s", fill=color, font=font)  # Replace with exact coordinates for t5
    draw.text((1400, 2600), f"T6B={t6}s", fill=color, font=font)  # Replace with exact coordinates for t6

    # Draw text on the image at specified positions
    draw.text((2320, 80), f"T7A={t7}s", fill=color, font=font)  # Replace with exact coordinates for t7
    draw.text((2450, 220), f"T8A={t8}s", fill=color, font=font)  # Replace with exact coordinates for t8
    draw.text((2750, 350), f"T9A={t9}s", fill=color, font=font)  # Replace with exact coordinates for t9
    draw.text((3100, 220), f"T10A={t10}s", fill=color, font=font)  # Replace with exact coordinates for t10

    draw.text((2320, 80+500), f"T7B={t7}s", fill=color, font=font)  # Replace with exact coordinates for t7
    draw.text((2450, 220+500), f"T8B={t8}s", fill=color, font=font)  # Replace with exact coordinates for t8
    draw.text((2750, 350+500), f"T9B={t9}s", fill=color, font=font)  # Replace with exact coordinates for t9
    draw.text((3100, 220+500), f"T10B={t10}s", fill=color, font=font)  # Replace with exact coordinates for t10

    draw.text((2320, 1090), f"T7C={t7}s", fill=color, font=font)  # Replace with exact coordinates for t7
    draw.text((2450, 1230), f"T8C={t8}s", fill=color, font=font)  # Replace with exact coordinates for t8
    draw.text((2750, 1360), f"T9C={t9}s", fill=color, font=font)  # Replace with exact coordinates for t9
    draw.text((3100, 1230), f"T10C={t10}s", fill=color, font=font)  # Replace with exact coordinates for t10

    draw.text((2320, 1590), f"T7D={t7}s", fill=color, font=font)  # Replace with exact coordinates for t7
    draw.text((2450, 1730), f"T8D={t8}s", fill=color, font=font)  # Replace with exact coordinates for t8
    draw.text((2750, 1860), f"T9D={t9}s", fill=color, font=font)  # Replace with exact coordinates for t9
    draw.text((3100, 1730), f"T10D={t10}s", fill=color, font=font)  # Replace with exact coordinates for t10

    draw.text((2320, 2140), f"T7E={t7}s", fill=color, font=font)  # Replace with exact coordinates for t7
    draw.text((2450, 2280), f"T8E={t8}s", fill=color, font=font)  # Replace with exact coordinates for t8
    draw.text((2750, 2410), f"T9E={t9}s", fill=color, font=font)  # Replace with exact coordinates for t9
    draw.text((3100, 2280), f"T10E={t10}s", fill=color, font=font)  # Replace with exact coordinates for t10

    draw.text((2320, 2600), f"T7F={t7}s", fill=color, font=font)  # Replace with exact coordinates for t7
    draw.text((2450, 2740), f"T8F={t8}s", fill=color, font=font)  # Replace with exact coordinates for t8
    draw.text((2750, 2870), f"T9F={t9}s", fill=color, font=font)  # Replace with exact coordinates for t9
    draw.text((3100, 2740), f"T10F={t10}s", fill=color, font=font)  # Replace with exact coordinates for t10

    return image


def generate_mill_data(num_mills=6, num_datasets=10):
    """
    Generate 10 datasets for 6 mills with selective components
    """
    # Define components and their colors
    components = [
        "A", "B", "C", "D", "E", "F", "G", "H", "I", "J",
        "ROM_S03_A", "ROM_S03_B", "ROM_S03_D",
        "ROM_S04_A", "ROM_S04_B", "ROM_S04_D",
        "ROM2_A", "ROM2_B", "ROM2_C", "ROM2_D"
    ]
    component_colors = [
        "#6AABF0", "#4B91E8", "#2F78D0", "#FFB685", "#F59A55", "#D47A30",
        "#89E0B5", "#68D69E", "#4CB47C", "#F7E08A", "#F4D35E", "#D1B342",
        "#7ADFF2", "#55D0E8", "#38A9BE", "#F2A5B3", "#EC7F9D", "#C86479",
        "#9B59B6", "#8E44AD"
    ]

    # Define weights, A, B, C, D have higher weights
    weights = {
        "A": 0.4, "B": 0.3, "C": 0.2, "D": 0.2,
        "E": 0.1, "F": 0.1, "G": 0.1, "H": 0.1, "I": 0.1, "J": 0.1,
        "ROM_S03_A": 0.05, "ROM_S03_B": 0.05, "ROM_S03_D": 0.05,
        "ROM_S04_A": 0.05, "ROM_S04_B": 0.05, "ROM_S04_D": 0.05,
        "ROM2_A": 0.05, "ROM2_B": 0.05, "ROM2_C": 0.05, "ROM2_D": 0.05
    }

    # Mills names
    mills = [f"Mill {i + 1}" for i in range(num_mills)]

    # Dataset to store all data
    mill_datasets = []
    random.seed(42)  # For reproducibility

    for _ in range(num_datasets):
        mill_data = {}
        for mill in mills:
            # Randomly select 5 to 10 components using weights
            num_components = random.randint(5, 10)
            selected_components = random.choices(
                components,
                weights=[weights[comp] for comp in components],
                k=num_components
            )

            # Randomly generate the values of these components so that the sum is close to 100
            values = [random.randint(10, 30) for _ in selected_components]
            total = sum(values)
            normalized_values = [round(v / total * 100, 2) for v in values]

            # Create a dictionary only for selected components
            mill_data[mill] = dict(zip(selected_components, normalized_values))

        mill_datasets.append(mill_data)

    return mill_datasets, components, component_colors, weights


def apply_weighted_average(datasets, refresh_interval, weights):
    interval_weights = {
        "5 mins": np.linspace(1, 0.5, 10),  # More weight to recent data
        "10 mins": np.linspace(1, 0.1, 10),  # Gradual decrease
        "30 mins": np.linspace(1, 0.01, 10),  # Sharp decrease
        "60 mins": np.ones(10)  # Uniform weight
    }

    normalized_weights = interval_weights[refresh_interval] / np.sum(interval_weights[refresh_interval])

    averaged_data = {}
    mills = list(datasets[0].keys())

    for mill in mills:
        # Collect all unique components across all datasets for this mill
        all_mill_components = set()
        for dataset in datasets:
            all_mill_components.update(dataset[mill].keys())

        mill_avg_data = {}
        for component in all_mill_components:
            component_values = [dataset[mill].get(component, 0) for dataset in datasets]
            weighted_avg = np.average(component_values, weights=normalized_weights)
            mill_avg_data[component] = round(weighted_avg, 2)

        averaged_data[mill] = mill_avg_data

    return averaged_data


def create_stacked_bar_chart(averaged_data, components, component_colors):
    mills = list(averaged_data.keys())

    fig = go.Figure()
    for i, component in enumerate(components):
        # Get the value of the current component in each mill
        y_values = [averaged_data[mill].get(component, 0) for mill in mills]
        if any(y_values):  # Only add trace if the component exists in at least one mill
            fig.add_trace(
                go.Bar(
                    x=mills,
                    y=y_values,
                    name=component,
                    marker_color=component_colors[i % len(component_colors)],
                    text=[component if v > 0 else "" for v in y_values],
                    textposition="inside",
                    customdata=[
                        f"<br>".join([f"{comp}: {val}%" for comp, val in averaged_data[mill].items() if val > 0])
                        for mill in mills
                    ],
                    hovertemplate="%{x}<br>%{customdata}"
                )
            )

    fig.update_layout(
        barmode="stack",
        title=f"Last updated: {datetime.datetime.now().strftime('%I:%M:%S %p %d-%m-%Y')}",
        xaxis_title="Mills",
        yaxis_title="Component Filling Level (%)",
        yaxis=dict(range=[0, 100]),
        showlegend=True,
        bargap=0.4,
        legend=dict(
            traceorder='normal',  # Legend order is from top to bottom
        ),
    )

    return fig


def load_data(file_path):
    """保留兼容性接口（已不使用文件路径，仅备用）"""
    try:
        df = pd.read_csv(file_path)
        df['Current Timestamp'] = pd.to_datetime(df['Current Timestamp'])
        return df
    except Exception as e:
        st.error(f"Error loading data: {e}")
        return None


def filter_data(df, start_date, end_date, exclude_columns, max_rows, exclude_empty,
                always_filled_cols=None):
    """
    always_filled_cols: columns that are always populated (e.g. cvr_name, belt_running)
    and should NOT count toward the "empty row" judgment.
    """
    if df is None:
        return None

    # Filter by date range
    mask = (df['Current Timestamp'] >= start_date) & (df['Current Timestamp'] <= end_date)
    filtered_df = df.loc[mask]

    # Remove excluded columns
    if exclude_columns:
        columns_to_show = [col for col in filtered_df.columns if col not in exclude_columns]
        filtered_df = filtered_df[columns_to_show]

    # Remove empty rows if selected
    if exclude_empty:
        skip = {'Current Timestamp'} | set(always_filled_cols or [])
        data_cols = [col for col in filtered_df.columns if col not in skip]
        if data_cols:
            filtered_df = filtered_df.dropna(subset=data_cols, how='all')

    # Limit number of rows
    if max_rows != 'Max' and len(filtered_df) > int(max_rows):
        filtered_df = filtered_df.head(int(max_rows))

    return filtered_df


def query_cvr_data(start_date, end_date):
    df = load_cvr_data_pg(st.session_state.selected_belt, start_date, end_date)
    if df is not None and not df.empty:
        st.session_state.current_df = df
        st.session_state.show_filters = True
        st.rerun()
    else:
        st.warning("No CVR data found for the selected date range.")


def query_tripper_data(start_date, end_date):
    df = load_tripper_data_pg(st.session_state.selected_tripper, start_date, end_date)
    if df is not None and not df.empty:
        st.session_state.current_tripper_df = df
        st.session_state.show_tripper_filters = True
        st.rerun()
    else:
        st.warning("No Tripper data found for the selected date range.")


def check_cvr_changes(selected_belt, start_date, end_date):
    if (selected_belt != st.session_state.previous_belt or
            (start_date, end_date) != st.session_state.previous_cvr_dates):
        st.session_state.show_filters = False
        st.session_state.current_df = None
    st.session_state.previous_belt = selected_belt
    st.session_state.previous_cvr_dates = (start_date, end_date)


def check_tripper_changes(selected_tripper, start_date, end_date):
    if (selected_tripper != st.session_state.previous_tripper or
            (start_date, end_date) != st.session_state.previous_tripper_dates):
        st.session_state.show_tripper_filters = False
        st.session_state.current_tripper_df = None
    st.session_state.previous_tripper = selected_tripper
    st.session_state.previous_tripper_dates = (start_date, end_date)


def load_stockpile_data(file_path):
    """保留兼容性接口（已不使用文件路径）"""
    try:
        df = pd.read_csv(file_path)
        df['Current Timestamp'] = pd.to_datetime(df['Current Timestamp'])
        return df
    except Exception as e:
        st.error(f"Error loading data: {e}")
        return None


def parse_layers_data(composition_str):
    """解析 silo_tracking 的 composition_tons/composition_pct 字段为 DataFrame"""
    try:
        if isinstance(composition_str, dict):
            comp_dict = composition_str
        elif isinstance(composition_str, str):
            comp_dict = json.loads(composition_str)
        else:
            return None
        if not comp_dict:
            return None
        df = pd.DataFrame([
            {'Source': k, 'Tons': round(float(v), 3)}
            for k, v in comp_dict.items()
        ])
        return df
    except:
        return None


def query_stockpile_data(start_date, end_date):
    silo_num = int(st.session_state.selected_stockpile.replace('stockpile', ''))
    df = load_silo_tracking_pg(silo_num, start_date, end_date)
    if df is not None and not df.empty:
        st.session_state.current_stockpile_df = df
        st.session_state.show_stockpile_filters = True
        st.rerun()
    else:
        st.warning("No stockpile data found for the selected date range.")


def check_stockpile_changes(selected_stockpile, start_date, end_date):
    if (selected_stockpile != st.session_state.previous_stockpile or
            (start_date, end_date) != st.session_state.previous_stockpile_dates):
        st.session_state.show_stockpile_filters = False
        st.session_state.current_stockpile_df = None
    st.session_state.previous_stockpile = selected_stockpile
    st.session_state.previous_stockpile_dates = (start_date, end_date)


def parse_mill_feed_layers(composition_str):
    """解析 mill_feed 的 silo*_composition_tons/pct 字段为 DataFrame"""
    return parse_layers_data(composition_str)


def load_mill_feed_data(file_path):
    """保留兼容性接口（已不使用文件路径）"""
    try:
        df = pd.read_csv(file_path)
        df['Current Timestamp'] = pd.to_datetime(df['Current Timestamp'])
        return df
    except Exception as e:
        st.error(f"Error loading data: {e}")
        return None


def query_mill_feed_data(start_date, end_date):
    mill_num = int(st.session_state.selected_mill_feed.replace('MillFeedCRV', ''))
    df = load_mill_feed_pg(mill_num, start_date, end_date)
    if df is not None and not df.empty:
        st.session_state.current_mill_feed_df = df
        st.session_state.show_mill_feed_filters = True
        st.session_state.show_stockpile_selector = False
        st.rerun()
    else:
        st.warning("No mill feed data found for the selected date range.")


def check_mill_feed_changes(selected_mill_feed, start_date, end_date):
    if (selected_mill_feed != st.session_state.previous_mill_feed or
            (start_date, end_date) != st.session_state.previous_mill_feed_dates):
        st.session_state.show_mill_feed_filters = False
        st.session_state.current_mill_feed_df = None
        st.session_state.show_stockpile_selector = False  # Reset stockpile selector visibility
    st.session_state.previous_mill_feed = selected_mill_feed
    st.session_state.previous_mill_feed_dates = (start_date, end_date)


def main():
    st.set_page_config(page_title='Citic Smart Tray App', initial_sidebar_state='expanded')
    st.logo(".\\imgs\\citic_logo.png", size="large")

    with open('config.yaml') as file:
        config = yaml.load(file, Loader=SafeLoader)

    authenticator = stauth.Authenticate(
        config['credentials'],
        config['cookie']['name'],
        config['cookie']['key'],
        config['cookie']['expiry_days'],
        config['preauthorized']
    )

    # Initialize authentication related state
    if 'authentication_status' not in st.session_state:
        st.session_state['authentication_status'] = None
    if 'name' not in st.session_state:
        st.session_state['name'] = None
    if 'username' not in st.session_state:
        st.session_state['username'] = None

    name, authentication_status, username = authenticator.login(fields={'Form name': 'Sino Iron - Ore Tracking and Prediction App', 'location': 'main'})

    if authentication_status:
        # Sidebar

        st.sidebar.toggle("Dark Mode")
        st.sidebar.header(":rainbow[Smart Ore Tracking App]")
        st.sidebar.image(".\\imgs\\dumpTruck.png", width=240)
        st.sidebar.subheader("***Next Gen Mine-to-Mill Intelligence***", divider='gray')

        app_chosen = st.sidebar.radio(
                        "Please Select App Function:",
                        ["Database Visualization",
                         "Ore Tracking Models and Visualisation",
                         "Ore Stockpile Filling Prediction",
                         "Mill Feed and Performance Forecast",
                         "AG Mill Performance Forecast",
                         "Configure Process Parameters",
                         "Configure Database Connection",]
        )


        ################################################
        if app_chosen == "Database Visualization":
            ### App 1 ##################
            st.subheader("Database Query System", divider="rainbow")

            # Database configuration settings
            DB_CONFIG = {
                "user": "postgres",
                "password": "postgres",
                "host": "localhost",
                "port": "5432"
            }

            # Mapping: category → { "db": actual_db, "tables": { label: (table, time_col) } }
            TABLE_MAPPING = {
                "Mine Operations": {
                    "db": "postgres",
                    "tables": {
                        "Truck Cycles":     ("truck_cycles",     "date"),
                        "Stockpile & Belt": ("realtime_data",    "date"),
                        "Mill Production":  ("production_lines", "时间"),
                    }
                },
                "Ore Tracking": {
                    "db": "postgres",
                    "tables": {
                        "CVR Tracking":     ("cvr_tracking",     "time"),
                        "Tripper Tracking": ("tripper_tracking", "time"),
                        "Silo Tracking":    ("silo_tracking",    "time"),
                        "Mill Feed":        ("mill_feed",        "time"),
                    }
                },
            }

            # Initialize Streamlit Session State
            if "query_result" not in st.session_state:
                st.session_state["query_result"] = None
            if "selected_columns" not in st.session_state:
                st.session_state["selected_columns"] = None

            # Use column layout for category and table selection
            col1, col2 = st.columns(2)
            with col1:
                db_choice = st.selectbox("Select Data Category", list(TABLE_MAPPING.keys()))
            with col2:
                table_choice = st.selectbox(
                    "Select Table",
                    list(TABLE_MAPPING[db_choice]["tables"].keys())
                )

            # Use column layout for date selection and interval
            col3, col4, col5 = st.columns(3)
            with col3:
                start_date = st.date_input("Select Start Date", datetime.datetime(2024, 1, 1))
            with col4:
                end_date = st.date_input("Select End Date", datetime.datetime.now())
            with col5:
                interval_choice = st.selectbox(
                    "Select Data Interval",
                    [1, 3, 5, 10, 30, 50, 100, 300, 500, 1000]
                )

            # Query button using full column width
            if st.button("Query Data", use_container_width=True):
                # Get corresponding table name, time field, and actual DB name
                table_name, time_field = TABLE_MAPPING[db_choice]["tables"][table_choice]
                actual_db = TABLE_MAPPING[db_choice]["db"]

                # Connect to database and query data
                try:
                    # Create a SQLAlchemy engine
                    engine = create_engine(
                        f'postgresql://{DB_CONFIG["user"]}:{DB_CONFIG["password"]}@{DB_CONFIG["host"]}:{DB_CONFIG["port"]}/{actual_db}'
                    )

                    # Build query SQL
                    query = f"""
                    SELECT * FROM {table_name}
                    WHERE {time_field} >= '{start_date}' AND {time_field} <= '{end_date}'
                    """
                    df = pd.read_sql(query, engine)  # Use engine to replace conn

                    # Store query results in session state
                    st.session_state["query_result"] = df
                    st.session_state["selected_columns"] = list(df.columns)

                    st.success("Query Successful!")
                except Exception as e:
                    st.error(f"Query Failed: {e}")

            # Display query results
            if st.session_state["query_result"] is not None:
                df = st.session_state["query_result"]

                # Column selection, default to all columns
                selected_columns = st.multiselect(
                    "Select Columns to Display",
                    options=df.columns,
                    default=df.columns
                )

                # Row limit options
                row_limit_options = [1000, 100, 500, 5000, 10000, 50000, 100000, "Custom Input", "All"]
                row_limit_choice = st.selectbox(
                    "Select Maximum Number of Rows to Display",
                    row_limit_options,
                    index=0
                )

                # Set number of rows based on user selection
                if row_limit_choice == "All":
                    max_rows = len(df)
                elif row_limit_choice == "Custom Input":
                    max_rows = st.number_input(
                        "Enter Maximum Number of Rows",
                        min_value=1,
                        max_value=len(df),
                        value=100,
                        step=1
                    )
                else:
                    max_rows = int(row_limit_choice)

                # Filtered data
                filtered_df = df[selected_columns].head(max_rows)

                # Display data
                st.dataframe(filtered_df)

                # Plotting
                st.markdown("**Data Visualization**")

                # Use column layout for X and Y axis selection
                col6, col7, col8 = st.columns(3)
                with col6:
                    x_axis = st.selectbox("Select X-Axis Column", options=selected_columns)
                with col7:
                    y_axis = st.selectbox("Select Y-Axis Column", options=selected_columns)
                with col8:
                    plot_type = st.selectbox(
                        "Select Plot Type",
                        ["Scatter", "Line", "Bar", "Box", "Violin"]
                    )

                if x_axis and y_axis:
                    # Create figure with custom theme
                    fig = go.Figure()

                    # Custom color palette
                    colors = {
                        'Scatter': '#1f77b4',  # Blue
                        'Line': '#2ca02c',  # Green
                        'Bar': '#ff7f0e',  # Orange
                        'Box': '#9467bd',  # Purple
                        'Violin': '#d62728'  # Red
                    }

                    # Data preparation and cleaning
                    x_data = filtered_df[x_axis]
                    y_data = filtered_df[y_axis]

                    # Convert data to numeric for trend line calculation
                    # datetime columns must be converted to Unix timestamps first
                    try:
                        if pd.api.types.is_datetime64_any_dtype(x_data):
                            x_numeric = x_data.astype('int64') / 1e9  # nanoseconds → seconds
                            x_is_datetime = True
                        else:
                            x_numeric = pd.to_numeric(x_data, errors='coerce')
                            x_is_datetime = False
                        y_numeric = pd.to_numeric(y_data, errors='coerce')

                        # Remove NaN values for trend line calculation
                        mask = ~(np.isnan(x_numeric) | np.isnan(y_numeric))
                        x_clean = x_numeric[mask]
                        y_clean = y_numeric[mask]
                        x_clean_dt = x_data[mask] if x_is_datetime else None
                    except Exception as e:
                        st.warning(f"Could not convert data to numeric format: {str(e)}")
                        x_clean = None
                        y_clean = None
                        x_clean_dt = None
                        x_is_datetime = False

                    # Add traces based on plot type
                    if plot_type == "Scatter":
                        fig.add_trace(go.Scatter(
                            x=x_data,
                            y=y_data,
                            mode='markers',
                            marker=dict(
                                size=8,
                                color=colors['Scatter'],
                                opacity=0.7,
                                line=dict(width=1, color='white'),
                                maxdisplayed=200
                            ),
                            name='Data Points'
                        ))

                        # Add trend line only if data is numeric and has sufficient valid points
                        if x_clean is not None and len(x_clean) > 1:
                            try:
                                valid_indices = np.isfinite(x_clean) & np.isfinite(y_clean)
                                if np.sum(valid_indices) > 1:
                                    x_valid = x_clean[valid_indices]
                                    y_valid = y_clean[valid_indices]

                                    sort_idx = np.argsort(x_valid)
                                    x_sorted = x_valid[sort_idx]
                                    y_sorted = y_valid[sort_idx]

                                    z = np.polyfit(x_sorted, y_sorted, 1)
                                    p = np.poly1d(z)

                                    # For datetime X, plot trend against original datetime values
                                    if x_is_datetime and x_clean_dt is not None:
                                        x_dt_valid = x_clean_dt[valid_indices]
                                        x_dt_sorted = x_dt_valid.iloc[sort_idx]
                                        fig.add_trace(go.Scatter(
                                            x=x_dt_sorted,
                                            y=p(x_sorted),
                                            mode='lines',
                                            name='Trend Line',
                                            line=dict(color='rgba(255, 0, 0, 0.5)', dash='dash')
                                        ))
                                    else:
                                        fig.add_trace(go.Scatter(
                                            x=x_sorted,
                                            y=p(x_sorted),
                                            mode='lines',
                                            name='Trend Line',
                                            line=dict(color='rgba(255, 0, 0, 0.5)', dash='dash')
                                        ))
                            except Exception:
                                st.info("Could not calculate trend line due to data characteristics")

                    elif plot_type == "Line":
                        fig.add_trace(go.Scatter(
                            x=x_data,
                            y=y_data,
                            mode='lines+markers',
                            line=dict(color=colors['Line'], width=2),
                            marker=dict(size=6),
                            name='Time Series'
                        ))

                    elif plot_type == "Bar":
                        fig.add_trace(go.Bar(
                            x=x_data,
                            y=y_data,
                            marker_color=colors['Bar'],
                            opacity=0.8,
                            name='Bar Chart'
                        ))

                    elif plot_type == "Box":
                        fig.add_trace(go.Box(
                            x=x_data,
                            y=y_data,
                            marker_color=colors['Box'],
                            name='Box Plot'
                        ))

                    elif plot_type == "Violin":
                        fig.add_trace(go.Violin(
                            x=x_data,
                            y=y_data,
                            fillcolor=colors['Violin'],
                            opacity=0.6,
                            name='Violin Plot'
                        ))

                    # Update layout with custom styling
                    fig.update_layout(
                        title=dict(
                            text=f"{plot_type} Plot: {x_axis} vs {y_axis}",
                            font=dict(size=14, color="#2c3e50"),
                            x=0.55,
                            y=0.85
                        ),
                        xaxis=dict(
                            title=dict(
                                text=x_axis,
                                font=dict(size=14)
                            ),
                            gridcolor='rgba(230, 230, 230, 0.6)',
                            showline=True,
                            linecolor='rgba(70, 70, 70, 0.3)',
                            linewidth=1,
                            tickfont=dict(size=12)
                        ),
                        yaxis=dict(
                            title=dict(
                                text=y_axis,
                                font=dict(size=14)
                            ),
                            gridcolor='rgba(230, 230, 230, 0.6)',
                            showline=True,
                            linecolor='rgba(70, 70, 70, 0.3)',
                            linewidth=1,
                            tickfont=dict(size=12)
                        ),
                        plot_bgcolor='white',
                        paper_bgcolor='white',
                        hovermode='closest',
                        showlegend=True,
                        legend=dict(
                            yanchor="top",
                            y=0.99,
                            xanchor="right",
                            x=0.99,
                            bgcolor='rgba(255, 255, 255, 0.8)',
                            bordercolor='rgba(70, 70, 70, 0.3)',
                            borderwidth=1
                        ),
                        margin=dict(l=80, r=80, t=100, b=80)
                    )

                    # Add grid
                    fig.update_xaxes(showgrid=True, gridwidth=1, gridcolor='rgba(230, 230, 230, 0.6)')
                    fig.update_yaxes(showgrid=True, gridwidth=1, gridcolor='rgba(230, 230, 230, 0.6)')

                    # Display the plot
                    st.plotly_chart(fig, use_container_width=True)

        elif app_chosen == "Ore Tracking Models and Visualisation":
            ### App 2 ##################
            st.subheader("Ore Tracking Models and Visualisation", divider="rainbow")
            st.markdown("**Please select a node to view:**")  # Added subtitle
            # Create tabs for different nodes
            tab1, tab2, tab3, tab4 = st.tabs([
                "N1: Crusher Belts",  # 破碎机下皮带
                "N2: Trippers",  # 给料小车
                "N3: Stockpiles",  # 原矿仓
                "N4: Mill Feed Belts"  # 磨机给料皮带
            ])

            # Initialize session state
            if 'current_df' not in st.session_state:
                st.session_state.current_df = None
            if 'show_filters' not in st.session_state:
                st.session_state.show_filters = False
            if 'current_tripper_df' not in st.session_state:
                st.session_state.current_tripper_df = None
            if 'show_tripper_filters' not in st.session_state:
                st.session_state.show_tripper_filters = False
            if 'date_range' not in st.session_state:
                current_date = datetime.datetime.now()
                st.session_state.date_range = {
                    'cvr_min': current_date,
                    'cvr_max': current_date,
                    'tripper_min': current_date,
                    'tripper_max': current_date
                }
            if 'previous_belt' not in st.session_state:
                st.session_state.previous_belt = None
            if 'previous_tripper' not in st.session_state:
                st.session_state.previous_tripper = None
            if 'previous_cvr_dates' not in st.session_state:
                st.session_state.previous_cvr_dates = (None, None)
            if 'previous_tripper_dates' not in st.session_state:
                st.session_state.previous_tripper_dates = (None, None)
            if 'current_stockpile_df' not in st.session_state:
                st.session_state.current_stockpile_df = None
            if 'show_stockpile_filters' not in st.session_state:
                st.session_state.show_stockpile_filters = False
            if 'selected_timestamp' not in st.session_state:
                st.session_state.selected_timestamp = None
            if 'previous_stockpile' not in st.session_state:
                st.session_state.previous_stockpile = None
            if 'previous_stockpile_dates' not in st.session_state:
                st.session_state.previous_stockpile_dates = (None, None)
            if 'current_mill_feed_df' not in st.session_state:
                st.session_state.current_mill_feed_df = None
            if 'show_mill_feed_filters' not in st.session_state:
                st.session_state.show_mill_feed_filters = False
            if 'previous_mill_feed' not in st.session_state:
                st.session_state.previous_mill_feed = None
            if 'previous_mill_feed_dates' not in st.session_state:
                st.session_state.previous_mill_feed_dates = (None, None)

            with tab1:
                # Belt selection
                belt_options = ['CVR111', 'CVR112', 'CVR113', 'CVR114']
                # Belt selection
                st.session_state.selected_belt = st.radio(
                    "Select Belt",  # 选择皮带
                    belt_options,
                    horizontal=True
                )

                # Date range selection - Always visible
                col1, col2 = st.columns(2)
                with col1:
                    start_date = st.date_input(
                        "Select Start Date",
                        datetime.date(2025, 4, 1)
                    )
                with col2:
                    end_date = st.date_input(
                        "Select End Date",
                        datetime.datetime.now()
                    )

                # Check for changes in selection or dates
                check_cvr_changes(st.session_state.selected_belt, start_date, end_date)

                # Query button
                if st.button("Query CVR Data", type="primary", use_container_width=True):
                    query_cvr_data(start_date, end_date)

                # Show filters and data only after query
                if st.session_state.show_filters and st.session_state.current_df is not None:
                    # Column selection
                    exclude_columns = st.multiselect(
                        "Select Columns to Hide",
                        [col for col in st.session_state.current_df.columns if col != 'Current Timestamp']
                    )

                    # Display controls
                    col3, col4 = st.columns(2)
                    with col3:
                        max_rows = st.selectbox(
                            "Maximum Rows to Display",
                            ['100', '500', '1000', '5000', 'Max'],
                            index=4
                        )
                    with col4:
                        exclude_empty = st.radio(
                            "Empty Rows Display",
                            ["Show All", "Hide Empty"],
                            index=0,
                            horizontal=True
                        )

                    # Filter and display data
                    # cvr_name and belt_running are always filled — exclude from empty-row check
                    filtered_df = filter_data(
                        st.session_state.current_df,
                        pd.Timestamp(start_date),
                        pd.Timestamp(end_date),
                        exclude_columns,
                        max_rows,
                        exclude_empty == "Hide Empty",
                        always_filled_cols=['cvr_name', 'belt_running']
                    )

                    if filtered_df is not None:
                        st.dataframe(filtered_df, use_container_width=True)

            with tab2:
                # Tripper selection
                tripper_options = ['Tripper1', 'Tripper2']
                st.session_state.selected_tripper = st.radio("Select Tripper", tripper_options, horizontal=True)

                # Date range selection - Always visible
                col1, col2 = st.columns(2)
                with col1:
                    start_date = st.date_input(
                        "Select Start Date",
                        datetime.date(2025, 4, 1),
                        key="tripper_start"
                    )
                with col2:
                    end_date = st.date_input(
                        "Select End Date",
                        datetime.datetime.now(),
                        key="tripper_end"
                    )

                # Check for changes in selection or dates
                check_tripper_changes(st.session_state.selected_tripper, start_date, end_date)

                # Query button
                if st.button("Query Tripper Data", type="primary", use_container_width=True):
                    query_tripper_data(start_date, end_date)

                # Show filters and data only after query
                if st.session_state.show_tripper_filters and st.session_state.current_tripper_df is not None:
                    # Column selection
                    exclude_columns = st.multiselect(
                        "Select Columns to Hide",
                        [col for col in st.session_state.current_tripper_df.columns if col != 'Current Timestamp'],
                        key="tripper_cols"
                    )

                    # Display controls
                    col3, col4 = st.columns(2)
                    with col3:
                        max_rows = st.selectbox(
                            "Maximum Rows to Display",
                            ['100', '500', '1000', '5000', 'Max'],
                            index=4,
                            key="tripper_rows"
                        )
                    with col4:
                        exclude_empty = st.radio(
                            "Empty Rows Display",
                            ["Show All", "Hide Empty"],
                            index=0,
                            horizontal=True,
                            key="tripper_empty"
                        )

                    # Filter and display data
                    filtered_df = filter_data(
                        st.session_state.current_tripper_df,
                        pd.Timestamp(start_date),
                        pd.Timestamp(end_date),
                        exclude_columns,
                        max_rows,
                        exclude_empty == "Hide Empty",
                        always_filled_cols=['tripper_name', 'belt_running']
                    )

                    if filtered_df is not None:
                        st.dataframe(filtered_df, use_container_width=True)

            with tab3:
                # Stockpile selection
                stockpile_options = [f'stockpile{i}' for i in range(1, 19)]
                st.session_state.selected_stockpile = st.selectbox("Select Stockpile", stockpile_options)

                # Date range selection
                col1, col2 = st.columns(2)
                with col1:
                    start_date = st.date_input(
                        "Select Start Date",
                        datetime.date(2025, 4, 1),
                        key="stockpile_start"
                    )
                with col2:
                    end_date = st.date_input(
                        "Select End Date",
                        datetime.datetime.now(),
                        key="stockpile_end"
                    )

                # Check for changes in selection or dates
                check_stockpile_changes(st.session_state.selected_stockpile, start_date, end_date)

                # Query button (full width)
                if st.button("Query Stockpile Data", type="primary", use_container_width=True):
                    query_stockpile_data(start_date, end_date)

                # Show filters and data only after query
                if st.session_state.show_stockpile_filters and st.session_state.current_stockpile_df is not None:
                    filtered_df = st.session_state.current_stockpile_df

                    # Read radio value from session state first so filtering runs before the widget renders
                    sp_empty_val = st.session_state.get("stockpile_empty", "Show All")
                    if sp_empty_val == "Hide Empty Layers":
                        filtered_df = filtered_df[
                            filtered_df['composition_tons'].apply(
                                lambda x: bool(x) if isinstance(x, dict) else (
                                    bool(json.loads(x)) if isinstance(x, str) and x not in ('{}', 'null', '') else False
                                )
                            )
                        ]

                    # Empty Layers radio (no label, horizontal — sits flush above the table)
                    st.radio(
                        "Empty Layers",
                        ["Hide Empty Layers", "Show All"],
                        index=0 if sp_empty_val == "Hide Empty Layers" else 1,
                        horizontal=True,
                        key="stockpile_empty",
                        label_visibility="collapsed"
                    )

                    # Summary table (filling level, mass, layers count)
                    display_cols = ['Current Timestamp', 'filling_level', 'mass', 'feed_amount',
                                    'discharge_amount', 'composition_number', 'layers_count']
                    display_cols = [c for c in display_cols if c in filtered_df.columns]
                    st.dataframe(filtered_df[display_cols], use_container_width=True)

                    # Below table: Date | Hour | Minute | Manual (single row)
                    # Parse manual HH:MM from previous run to sync selectors
                    _sp_manual_raw = st.session_state.get("stockpile_ts_manual", "").strip()
                    _sp_mhh = _sp_mmm = None
                    if _sp_manual_raw and ":" in _sp_manual_raw:
                        _sp_parts = _sp_manual_raw.split(":")
                        if len(_sp_parts) >= 2:
                            _sp_mhh = _sp_parts[0].zfill(2)
                            _sp_mmm = _sp_parts[1][:2].zfill(2)

                    ts_series = filtered_df['Current Timestamp']
                    available_dates = sorted(ts_series.dt.date.unique())
                    r1c1, r1c2, r1c3, r1c4 = st.columns(4)
                    with r1c1:
                        sel_date = st.selectbox(
                            "Select Date",
                            available_dates,
                            key="stockpile_ts_date"
                        )
                    ts_for_date = ts_series[ts_series.dt.date == sel_date]
                    sp_hours = sorted(ts_for_date.dt.strftime('%H').unique().tolist())
                    if _sp_mhh and _sp_mhh in sp_hours:
                        st.session_state["stockpile_ts_hour"] = _sp_mhh
                    with r1c2:
                        sel_hour = st.selectbox(
                            "Select Hour",
                            sp_hours,
                            index=len(sp_hours) - 1 if sp_hours else 0,
                            key="stockpile_ts_hour"
                        )
                    sp_mins = sorted(
                        ts_for_date[ts_for_date.dt.strftime('%H') == sel_hour]
                        .dt.strftime('%M').unique().tolist()
                    )
                    if _sp_mmm and _sp_mmm in sp_mins:
                        st.session_state["stockpile_ts_min"] = _sp_mmm
                    with r1c3:
                        sel_min = st.selectbox(
                            "Select Minute",
                            sp_mins,
                            index=len(sp_mins) - 1 if sp_mins else 0,
                            key="stockpile_ts_min"
                        )
                    with r1c4:
                        manual_ts = st.text_input(
                            "Or enter manually",
                            value="",
                            key="stockpile_ts_manual",
                            placeholder="HH:MM"
                        )

                    if _sp_mhh and _sp_mmm:
                        selected_timestamp = f"{sel_date} {_sp_mhh}:{_sp_mmm}:00"
                    else:
                        selected_timestamp = f"{sel_date} {sel_hour}:{sel_min}:00"

                    if st.button("View Composition Detail", type="primary", use_container_width=True, key="stockpile_view_button"):
                        row = filtered_df[
                            filtered_df['Current Timestamp'].dt.strftime('%Y-%m-%d %H:%M:%S') == selected_timestamp
                        ]
                        if not row.empty:
                            col_tons, col_pct = st.columns(2)
                            with col_tons:
                                st.markdown("**Composition (Tons)**")
                                comp_tons_df = parse_layers_data(row.iloc[0].get('composition_tons'))
                                if comp_tons_df is not None:
                                    st.dataframe(comp_tons_df, use_container_width=True)
                                else:
                                    st.info("No composition data")
                            with col_pct:
                                st.markdown("**Composition (%)**")
                                comp_pct_df = parse_layers_data(row.iloc[0].get('composition_pct'))
                                if comp_pct_df is not None:
                                    comp_pct_df = comp_pct_df.rename(columns={'Tons': 'Pct'})
                                    st.dataframe(comp_pct_df, use_container_width=True)
                                else:
                                    st.info("No percentage data")

            with tab4:
                # Add new session state variables
                if 'show_stockpile_selector' not in st.session_state:
                    st.session_state.show_stockpile_selector = False
                if 'current_layers_df' not in st.session_state:
                    st.session_state.current_layers_df = None
                if 'selected_timestamp_data' not in st.session_state:
                    st.session_state.selected_timestamp_data = None

                # Mill Feed Belt selection
                mill_feed_options = [f'MillFeedCRV{i}' for i in range(1, 7)]
                st.session_state.selected_mill_feed = st.selectbox("Select Mill Feed Belt", mill_feed_options)

                # Date range selection
                col1, col2 = st.columns(2)
                with col1:
                    start_date = st.date_input(
                        "Select Start Date",
                        datetime.date(2025, 4, 1),
                        key="mill_feed_start"
                    )
                with col2:
                    end_date = st.date_input(
                        "Select End Date",
                        datetime.datetime.now(),
                        key="mill_feed_end"
                    )

                # Check for changes in selection or dates
                check_mill_feed_changes(st.session_state.selected_mill_feed, start_date, end_date)

                # Query button (full width)
                if st.button("Query Mill Feed Data", type="primary", use_container_width=True,
                             key="mill_feed_query_button"):
                    query_mill_feed_data(start_date, end_date)

                # Show filters and data only after query
                if st.session_state.show_mill_feed_filters and st.session_state.current_mill_feed_df is not None:
                    filtered_df = st.session_state.current_mill_feed_df

                    # Read radio value from session state first so filtering runs before widget renders
                    mf_empty_val = st.session_state.get("mill_feed_empty", "Show All")
                    if mf_empty_val == "Hide Empty Layers":
                        filtered_df = filtered_df[
                            filtered_df['mill_composition_tons'].apply(
                                lambda x: bool(x) if isinstance(x, dict) else (
                                    bool(json.loads(x)) if isinstance(x, str) and x not in ('{}', 'null', '') else False
                                )
                            )
                        ]

                    # Empty Composition radio (no label, horizontal — sits flush above the table)
                    st.radio(
                        "Empty Composition",
                        ["Hide Empty Layers", "Show All"],
                        index=0 if mf_empty_val == "Hide Empty Layers" else 1,
                        horizontal=True,
                        key="mill_feed_empty",
                        label_visibility="collapsed"
                    )

                    # Summary table
                    display_cols = ['Current Timestamp', 'mill_throughput', 'calculated_throughput',
                                    'mill_composition_numbers', 'silo1_num', 'silo1_discharge',
                                    'silo2_num', 'silo2_discharge', 'silo3_num', 'silo3_discharge']
                    display_cols = [c for c in display_cols if c in filtered_df.columns]
                    st.dataframe(filtered_df[display_cols], use_container_width=True)

                    # Below table: Date | Hour | Minute | Manual (single row)
                    # Parse manual HH:MM from previous run to sync selectors
                    _mf_manual_raw = st.session_state.get("mill_feed_ts_manual", "").strip()
                    _mf_mhh = _mf_mmm = None
                    if _mf_manual_raw and ":" in _mf_manual_raw:
                        _mf_parts = _mf_manual_raw.split(":")
                        if len(_mf_parts) >= 2:
                            _mf_mhh = _mf_parts[0].zfill(2)
                            _mf_mmm = _mf_parts[1][:2].zfill(2)

                    ts_series_mf = filtered_df['Current Timestamp']
                    available_dates_mf = sorted(ts_series_mf.dt.date.unique())
                    mf_r1c1, mf_r1c2, mf_r1c3, mf_r1c4 = st.columns(4)
                    with mf_r1c1:
                        sel_date_mf = st.selectbox(
                            "Select Date",
                            available_dates_mf,
                            key="mill_feed_ts_date"
                        )
                    ts_for_date_mf = ts_series_mf[ts_series_mf.dt.date == sel_date_mf]
                    mf_hours = sorted(ts_for_date_mf.dt.strftime('%H').unique().tolist())
                    if _mf_mhh and _mf_mhh in mf_hours:
                        st.session_state["mill_feed_ts_hour"] = _mf_mhh
                    with mf_r1c2:
                        sel_hour_mf = st.selectbox(
                            "Select Hour",
                            mf_hours,
                            index=len(mf_hours) - 1 if mf_hours else 0,
                            key="mill_feed_ts_hour"
                        )
                    mf_mins = sorted(
                        ts_for_date_mf[ts_for_date_mf.dt.strftime('%H') == sel_hour_mf]
                        .dt.strftime('%M').unique().tolist()
                    )
                    if _mf_mmm and _mf_mmm in mf_mins:
                        st.session_state["mill_feed_ts_min"] = _mf_mmm
                    with mf_r1c3:
                        sel_min_mf = st.selectbox(
                            "Select Minute",
                            mf_mins,
                            index=len(mf_mins) - 1 if mf_mins else 0,
                            key="mill_feed_ts_min"
                        )
                    with mf_r1c4:
                        manual_ts_mf = st.text_input(
                            "Or enter manually",
                            value="",
                            key="mill_feed_ts_manual",
                            placeholder="HH:MM"
                        )

                    if _mf_mhh and _mf_mmm:
                        selected_timestamp = f"{sel_date_mf} {_mf_mhh}:{_mf_mmm}:00"
                    else:
                        selected_timestamp = f"{sel_date_mf} {sel_hour_mf}:{sel_min_mf}:00"

                    # View button — guard against no matching row
                    if st.button("View Composition Detail", type="primary", use_container_width=True, key="mill_feed_view_button"):
                        matched = filtered_df[
                            filtered_df['Current Timestamp'].dt.strftime('%Y-%m-%d %H:%M:%S') == selected_timestamp
                        ]
                        if not matched.empty:
                            st.session_state.show_stockpile_selector = True
                            st.session_state.selected_timestamp_data = matched.iloc[0]
                        else:
                            st.warning(f"No data found for timestamp: {selected_timestamp}")

                    # Show silo composition after view button is clicked
                    if st.session_state.show_stockpile_selector and st.session_state.selected_timestamp_data is not None:
                        selected_silo_slot = st.radio(
                            "Select Silo Composition",
                            ["Mill Overall", "Silo 1", "Silo 2", "Silo 3"],
                            horizontal=True,
                            key="mill_feed_stockpile_select",
                        )

                        col_map = {
                            "Mill Overall": ("mill_composition_tons", "mill_composition_pct"),
                            "Silo 1": ("silo1_composition_tons", "silo1_composition_pct"),
                            "Silo 2": ("silo2_composition_tons", "silo2_composition_pct"),
                            "Silo 3": ("silo3_composition_tons", "silo3_composition_pct"),
                        }
                        col_tons_key, col_pct_key = col_map[selected_silo_slot]
                        row_data = st.session_state.selected_timestamp_data

                        col_t, col_p = st.columns(2)
                        with col_t:
                            st.markdown(f"**{selected_silo_slot} Composition (Tons)**")
                            df_t = parse_mill_feed_layers(row_data.get(col_tons_key))
                            if df_t is not None:
                                st.dataframe(df_t, use_container_width=True)
                            else:
                                st.info("No data")
                        with col_p:
                            st.markdown(f"**{selected_silo_slot} Composition (%)**")
                            df_p = parse_mill_feed_layers(row_data.get(col_pct_key))
                            if df_p is not None:
                                df_p = df_p.rename(columns={'Tons': 'Pct'})
                                st.dataframe(df_p, use_container_width=True)
                            else:
                                st.info("No data")



        elif app_chosen == "Ore Stockpile Filling Prediction":
            ### App 3 ##################
            st.subheader("RoM Ore Stockpile Reconcilliation", divider="rainbow")
            ##################################################

            # Add option to choose between tons or percentage
            display_option = st.selectbox("Display Data As:", ["Tonnage", "Percentage"])

            cl0 = st.columns([1,1], gap="small", vertical_alignment="center")
            with cl0[0]:
                if st.button("Confirm", use_container_width=True, key="real_time_confirm"):
                    st.session_state.real_time_confirmed = True
            with cl0[-1]:
                st.button("Refresh Now", use_container_width=True)

            # 从 silo_tracking 查询每个料仓的最新数据
            @st.cache_data(ttl=60 * 5)
            def load_silo_data(_engine_str):
                """查询 silo_tracking 每个料仓的最新一条记录"""
                silo_data = {}
                try:
                    engine = create_engine(_engine_str)
                    query = text("""
                        SELECT DISTINCT ON (silo_num)
                            silo_num, time, filling_level,
                            composition_tons, composition_pct
                        FROM silo_tracking
                        ORDER BY silo_num, time DESC
                    """)
                    with engine.connect() as conn:
                        df = pd.read_sql(query, conn)
                    for _, row in df.iterrows():
                        i = int(row['silo_num'])

                        def parse_json_col(val):
                            if isinstance(val, dict):
                                return val
                            if isinstance(val, str) and val not in ('{}', 'null', ''):
                                try:
                                    return json.loads(val)
                                except Exception:
                                    return {}
                            return {}

                        silo_data[f"S{i}"] = {
                            'time': row['time'],
                            'level': float(row.get('filling_level') or 0),
                            'tonnage': parse_json_col(row.get('composition_tons')),
                            'percentage': parse_json_col(row.get('composition_pct')),
                        }
                except Exception as e:
                    st.error(f"Error querying silo_tracking: {e}")
                return silo_data

            # Load data
            try:
                silo_data = load_silo_data(DB_CONNECTION)
            except Exception as e:
                st.error(f"Error loading data: {e}")
                silo_data = {}

            if not silo_data:
                st.warning("No silo data available. Please ensure silo_tracking table has data.")

            # Derive available date range from silo_tracking for historical section
            @st.cache_data(ttl=300)
            def get_silo_date_range(_engine_str):
                try:
                    engine = create_engine(_engine_str)
                    with engine.connect() as conn:
                        row = conn.execute(text(
                            "SELECT MIN(time)::date AS min_d, MAX(time)::date AS max_d FROM silo_tracking"
                        )).fetchone()
                    if row and row[0]:
                        return row[0], row[1]
                except Exception:
                    pass
                return datetime.date.today() - datetime.timedelta(days=30), datetime.date.today()

            silo_min_date, silo_max_date = get_silo_date_range(DB_CONNECTION)
            sorted_dates = [silo_max_date]  # compatibility reference for historical section

            # Extract all unique component types across all silos
            all_components = set()
            for silo in silo_data.values():
                if display_option == "Tonnage":
                    components = silo.get('tonnage', {})
                else:  # Percentage
                    components = silo.get('percentage', {})

                for component in components:
                    all_components.add(component)

            # Convert to sorted list for consistent ordering
            all_components = sorted(list(all_components))

            # Create color mapping for components
            # Generate a list of colors
            colorscale = [
                "#6AABF0", "#4B91E8", "#2F78D0", "#FFB685", "#F59A55", "#D47A30",
                "#89E0B5", "#68D69E", "#4CB47C", "#F7E08A", "#F4D35E", "#D1B342",
                "#7ADFF2", "#55D0E8", "#38A9BE", "#F2A5B3", "#EC7F9D", "#C86479",
                "#9B59B6", "#8E44AD", "#745399", "#D35400", "#2ECC71", "#3498DB",
                "#E74C3C", "#34495E", "#1ABC9C", "#F1C40F", "#7F8C8D", "#27AE60"
            ]

            # Map components to colors (cycling through the color list if needed)
            component_colors = {}
            for i, component in enumerate(all_components):
                component_colors[component] = colorscale[i % len(colorscale)]

            # Extract data for the chart
            categories = list(silo_data.keys())  # S1, S2, etc.

            # Create hover texts
            hover_texts = []
            for silo_id in categories:
                if silo_id in silo_data:
                    if display_option == "Tonnage":
                        components_data = silo_data[silo_id].get('tonnage', {})
                    else:  # Percentage
                        components_data = silo_data[silo_id].get('percentage', {})

                    hover_text = f"<b>{silo_id}</b><br>"
                    hover_text += "<br>".join(
                        [f"{comp}: {round(val, 2)} {'tons' if display_option == 'Tonnage' else '%'}"
                         for comp, val in components_data.items()])
                    hover_texts.append(hover_text)
                else:
                    hover_texts.append(f"<b>{silo_id}</b><br>No data")

            # Create a stacked bar chart
            fig = go.Figure()

            for component in all_components:
                # Get the value of the current component in each silo
                y_values = []
                for silo_id in categories:
                    if silo_id in silo_data:
                        if display_option == "Tonnage":
                            components_data = silo_data[silo_id].get('tonnage', {})
                        else:  # Percentage
                            components_data = silo_data[silo_id].get('percentage', {})

                        y_values.append(components_data.get(component, 0))
                    else:
                        y_values.append(0)

                fig.add_trace(
                    go.Bar(
                        x=categories,
                        y=y_values,
                        name=component,
                        marker_color=component_colors.get(component, "#CCCCCC"),
                        text=[component if v > 0 else "" for v in y_values],
                        textposition="inside",
                        hoverinfo="none"  # We'll use a custom hovertemplate
                    )
                )

            # Update bar layout to use specific hover data
            for i, silo_id in enumerate(categories):
                fig.add_trace(
                    go.Scatter(
                        x=[silo_id],
                        y=[0],  # Position at the bottom
                        mode="markers",
                        marker=dict(size=0, opacity=0),  # Make the marker invisible
                        hoverinfo="text",
                        hovertext=hover_texts[i],
                        showlegend=False
                    )
                )

            # Update layout
            fig.update_layout(
                barmode="stack",
                title=f"Real-time Stockpile Status - Last updated: {max((v['time'] for v in silo_data.values() if v.get('time') is not None), default=datetime.datetime.now()).strftime('%I:%M:%S %p %d-%m-%Y')}",
                xaxis_title="Stockpile Silo",
                yaxis_title=f"Stockpile Composition ({'Tons' if display_option == 'Tonnage' else 'Percentage'})",
                showlegend=True,
                legend=dict(
                    traceorder='normal',
                ),
                bargap=0.05,
                hovermode="closest"
            )

            # If showing percentage, set y-axis range to 0-100
            if display_option == "Percentage":
                fig.update_layout(yaxis=dict(range=[0, 100]))

            st.plotly_chart(fig, use_container_width=True)

            # Historical Data Tracking Section
            st.subheader("Historical Data Tracking", divider="rainbow")

            # 从 silo_tracking 获取指定日期+料仓的可用时间点
            @st.cache_data
            def get_available_times(date, silo_number):
                try:
                    engine = create_engine(DB_CONNECTION)
                    query = text("""
                        SELECT DISTINCT to_char(time, 'HH24:MI') AS hm
                        FROM silo_tracking
                        WHERE silo_num = :num
                          AND time::date = :d
                        ORDER BY hm
                    """)
                    with engine.connect() as conn:
                        df = pd.read_sql(query, conn, params={
                            'num': int(silo_number), 'd': str(date)
                        })
                    return df['hm'].tolist() if not df.empty else []
                except Exception:
                    return []

            # 从 silo_tracking 加载7天历史数据（指定料仓、时间点、聚合窗口）
            @st.cache_data
            def load_historical_data(end_date, end_time, silo_number, interval_mins):
                hour, minute = map(int, end_time.split(':'))
                end_datetime = datetime.datetime.combine(end_date, datetime.time(hour, minute))
                start_date = end_date - datetime.timedelta(days=6)
                historical_data = {}

                def parse_jb(val):
                    if isinstance(val, dict):
                        return val
                    if isinstance(val, str) and val not in ('{}', 'null', ''):
                        try:
                            return json.loads(val)
                        except Exception:
                            return {}
                    return {}

                try:
                    engine = create_engine(DB_CONNECTION)
                    interval_minutes = 1 if interval_mins == "1 mins" else int(interval_mins.split()[0])
                    query = text("""
                        SELECT time, composition_tons, composition_pct
                        FROM silo_tracking
                        WHERE silo_num = :num
                          AND time >= :start AND time <= :end
                        ORDER BY time
                    """)
                    with engine.connect() as conn:
                        df = pd.read_sql(query, conn, params={
                            'num': int(silo_number),
                            'start': datetime.datetime.combine(start_date, datetime.time(0, 0)),
                            'end': end_datetime,
                        })
                    df['time'] = pd.to_datetime(df['time'])

                    current_date = start_date
                    while current_date <= end_date:
                        target_time = datetime.datetime.combine(current_date, datetime.time(hour, minute))
                        window_start = target_time - datetime.timedelta(minutes=interval_minutes)
                        window = df[(df['time'] >= window_start) & (df['time'] <= target_time)]

                        if not window.empty:
                            all_tonnage = {}
                            all_percentage = {}
                            for _, row in window.iterrows():
                                for comp, tons in parse_jb(row['composition_tons']).items():
                                    all_tonnage[comp] = all_tonnage.get(comp, 0) + tons
                                for comp, pct in parse_jb(row['composition_pct']).items():
                                    all_percentage.setdefault(comp, []).append(pct)

                            avg_percentage = {c: sum(v) / len(v) for c, v in all_percentage.items()}
                            historical_data[current_date] = {
                                'time': target_time,
                                'tonnage': all_tonnage,
                                'percentage': avg_percentage,
                            }
                        current_date += datetime.timedelta(days=1)
                except Exception as e:
                    st.warning(f"Error loading historical data: {e}")

                return historical_data


            ####   **************         进度条会被覆盖       **************
            #  @st.cache_data
            # def load_historical_data(end_date, end_time, silo_number, interval_mins):
            #     # Parse the end time
            #     hour, minute = map(int, end_time.split(':'))
            #
            #     # Create a datetime object for the end date and time
            #     end_datetime = datetime.datetime.combine(end_date, datetime.time(hour, minute))
            #
            #     # Calculate start date (7 days before)
            #     start_date = end_date - datetime.timedelta(days=6)
            #
            #     # Dictionary to store data for each date
            #     historical_data = {}
            #
            #     # Create a progress bar
            #     total_days = (end_date - start_date).days + 1
            #     progress_bar = st.progress(0)
            #     status_text = st.empty()
            #
            #     # Process each date
            #     current_date = start_date
            #     day_counter = 0
            #
            #     while current_date <= end_date:
            #         # Update progress
            #         day_counter += 1
            #         progress = day_counter / total_days
            #         progress_bar.progress(progress)
            #         status_text.text(
            #             f"Processing data for {current_date.strftime('%Y-%m-%d')} ({day_counter}/{total_days})")
            #
            #         date_str = current_date.strftime("%Y%m%d")
            #         file_path = os.path.join(data_dir, f"Silos_Data_{date_str}.xlsx")
            #
            # Row 1: Date | Silo
            col1, col2 = st.columns(2)
            with col1:
                selected_date = st.date_input(
                    "Select Date",
                    value=silo_max_date,
                    min_value=silo_min_date,
                    max_value=silo_max_date
                )
            with col2:
                selected_silo = st.selectbox(
                    "Select Silo",
                    options=[f"{i}" for i in range(1, 19)]
                )

            # Build hour→minutes mapping from available times
            available_times = get_available_times(selected_date, selected_silo)
            if available_times:
                from collections import defaultdict as _dd
                hour_to_mins = _dd(list)
                for t in available_times:
                    hh, mm = t.split(":")
                    hour_to_mins[hh].append(mm)
                available_hours = sorted(hour_to_mins.keys())
            else:
                hour_to_mins = {}
                available_hours = []

            # Row 2: Hour | Minute | Or enter manually
            col_h, col_m, col_manual = st.columns(3)
            with col_h:
                if available_hours:
                    sel_hour = st.selectbox(
                        "Hour",
                        options=available_hours,
                        index=len(available_hours) - 1,
                        key="hist_hour"
                    )
                else:
                    sel_hour = "00"
                    st.selectbox("Hour", ["--"], key="hist_hour", disabled=True)

            with col_m:
                avail_mins = sorted(hour_to_mins.get(sel_hour, []))
                if avail_mins:
                    sel_min = st.selectbox(
                        "Minute",
                        options=avail_mins,
                        index=len(avail_mins) - 1,
                        key="hist_minute"
                    )
                else:
                    sel_min = "00"
                    st.selectbox("Minute", ["--"], key="hist_minute", disabled=True)

            with col_manual:
                manual_hist_time = st.text_input(
                    "Or enter manually (HH:MM)",
                    value="",
                    key="hist_manual_time",
                    placeholder="e.g. 08:30"
                )

            if manual_hist_time.strip():
                selected_time = manual_hist_time.strip()
            elif available_times:
                selected_time = f"{sel_hour}:{sel_min}"
            else:
                selected_time = "00:00"

            # Button to confirm and load historical data
            load_historical = st.button("Confirm", use_container_width=True, key="hist_confirm")  ## type="primary",

            # Initialize session state for historical data if not already done
            if "historical_data" not in st.session_state:
                st.session_state.historical_data = None
                st.session_state.hist_confirmed = False
                st.session_state.hist_params = {
                    "date": None,
                    "time": None,
                    "silo": None
                }

            # Load historical data if confirmed
            if load_historical:
                st.session_state.hist_confirmed = True
                st.session_state.hist_params = {
                    "date": selected_date,
                    "time": selected_time,
                    "silo": int(selected_silo)
                }

                # Default interval for initial load
                default_interval = "1 mins"

                st.session_state.historical_data = load_historical_data(
                    selected_date,
                    selected_time,
                    int(selected_silo),
                    default_interval
                )

            # Display historical data controls and visualization if confirmed
            if st.session_state.hist_confirmed and st.session_state.historical_data:
                # Historical interval selection - AFTER confirmation
                # st.write("### Historical Data Visualization")
                hist_refresh_interval = st.selectbox(
                    "Historical Data Interval",
                    ["1 mins", "5 mins", "10 mins", "30 mins", "60 mins"],
                    key="hist_interval_display"
                )

                # Check if interval changed
                if "last_interval" not in st.session_state:
                    st.session_state.last_interval = hist_refresh_interval

                # Reload data if interval changed
                if st.session_state.last_interval != hist_refresh_interval:
                    st.session_state.historical_data = load_historical_data(
                        st.session_state.hist_params["date"],
                        st.session_state.hist_params["time"],
                        st.session_state.hist_params["silo"],
                        hist_refresh_interval
                    )
                    st.session_state.last_interval = hist_refresh_interval

                historical_data = st.session_state.historical_data

                if historical_data:
                    # Create a historical stacked bar chart
                    hist_fig = go.Figure()

                    # Extract all unique components
                    hist_components = set()
                    for date_data in historical_data.values():
                        if display_option == "Tonnage":
                            components = date_data.get('tonnage', {})
                        else:  # Percentage
                            components = date_data.get('percentage', {})

                        for component in components:
                            hist_components.add(component)

                    # Convert to sorted list
                    hist_components = sorted(list(hist_components))

                    # Map components to colors
                    hist_colors = {}
                    for i, component in enumerate(hist_components):
                        hist_colors[component] = colorscale[i % len(colorscale)]

                    # Sort dates
                    sorted_hist_dates = sorted(historical_data.keys())

                    # Create x-axis labels
                    x_labels = [date.strftime("%Y-%m-%d") for date in sorted_hist_dates]

                    # Create hover texts
                    hist_hover_texts = []
                    for date in sorted_hist_dates:
                        date_data = historical_data.get(date, {})

                        if display_option == "Tonnage":
                            components_data = date_data.get('tonnage', {})
                        else:  # Percentage
                            components_data = date_data.get('percentage', {})

                        hover_text = f"<b>{date.strftime('%Y-%m-%d')}</b><br>"
                        hover_text += "<br>".join(
                            [f"{comp}: {round(val, 2)} {'tons' if display_option == 'Tonnage' else '%'}"
                             for comp, val in components_data.items()])
                        hist_hover_texts.append(hover_text)

                    # Add traces for each component
                    for component in hist_components:
                        # Get values for this component across dates
                        y_values = []
                        for date in sorted_hist_dates:
                            date_data = historical_data.get(date, {})

                            if display_option == "Tonnage":
                                components_data = date_data.get('tonnage', {})
                            else:  # Percentage
                                components_data = date_data.get('percentage', {})

                            y_values.append(components_data.get(component, 0))

                        hist_fig.add_trace(
                            go.Bar(
                                x=x_labels,
                                y=y_values,
                                name=component,
                                marker_color=hist_colors.get(component, "#CCCCCC"),
                                text=[component if v > 0 else "" for v in y_values],
                                textposition="inside",
                                hoverinfo="none"
                            )
                        )

                    # Add invisible traces for hover
                    for i, date_str in enumerate(x_labels):
                        hist_fig.add_trace(
                            go.Scatter(
                                x=[date_str],
                                y=[0],
                                mode="markers",
                                marker=dict(size=0, opacity=0),
                                hoverinfo="text",
                                hovertext=hist_hover_texts[i],
                                showlegend=False
                            )
                        )

                    # Update layout
                    hist_fig.update_layout(
                        barmode="stack",
                        title=f"Historical Data for Silo {st.session_state.hist_params['silo']} at {st.session_state.hist_params['time']} ({hist_refresh_interval} interval)",
                        xaxis_title="Date",
                        yaxis_title=f"Composition ({'Tons' if display_option == 'Tonnage' else 'Percentage'})",
                        showlegend=True,
                        legend=dict(
                            traceorder='normal',
                        ),
                        bargap=0.1,
                        hovermode="closest"
                    )

                    # If showing percentage, set y-axis range to 0-100
                    if display_option == "Percentage":
                        hist_fig.update_layout(yaxis=dict(range=[0, 100]))

                    st.plotly_chart(hist_fig, use_container_width=True)

                    # # Show summary of the data
                    # with st.expander("View Historical Data Details"):
                    #     for date in sorted_hist_dates:
                    #         date_data = historical_data.get(date, {})
                    #         st.write(f"**{date.strftime('%Y-%m-%d')} at {st.session_state.hist_params['time']}**")
                    #
                    #         if display_option == "Tonnage":
                    #             components_data = date_data.get('tonnage', {})
                    #             st.write(f"Total tonnage: {sum(components_data.values()):.2f} tons")
                    #         else:
                    #             components_data = date_data.get('percentage', {})
                    #
                    #         # Create a small table of components
                    #         component_df = pd.DataFrame({
                    #             'Component': list(components_data.keys()),
                    #             f'{"Tons" if display_option == "Tonnage" else "Percentage"}': [
                    #                 f"{components_data[comp]:.2f}" for comp in components_data.keys()
                    #             ]
                    #         })
                    #         st.dataframe(component_df)
                    #         st.write("---")
                else:
                    st.warning("No historical data available for the selected parameters")




        elif app_chosen == "Mill Feed and Performance Forecast":
            ### App 4 ##################
            st.subheader("Mill Feed Prediction", divider="rainbow")
            ##################################################
            # Data refresh interval selection for both visualizations
            refresh_interval = st.selectbox(
                "Please Select Data Refresh Interval",
                ["1 min", "5 mins", "10 mins", "30 mins", "60 mins"],
                index=0
            )

            cl1 = st.columns([1, 1, 1, 1], gap="medium", vertical_alignment="center")
            with cl1[0]:
                st.button("Confirm", key="confirm_mill_feed", use_container_width=True)
            with cl1[-1]:
                st.button("Refresh Now", key="refresh_mill_feed", use_container_width=True)

            # Function to aggregate data based on refresh interval
            def aggregate_data(df, interval):
                # Convert refresh interval to minutes
                if interval == "1 min":
                    return df  # No aggregation needed

                minutes = int(interval.split()[0])

                # Create a copy of the dataframe
                aggregated_df = df.copy()

                # Convert time column to datetime if it's not already
                if not pd.api.types.is_datetime64_any_dtype(aggregated_df.iloc[:, 0]):
                    aggregated_df.iloc[:, 0] = pd.to_datetime(aggregated_df.iloc[:, 0])

                # Group by time intervals and calculate mean
                aggregated_df.set_index(aggregated_df.columns[0], inplace=True)
                resampled = aggregated_df.resample(f'{minutes}T').mean().reset_index()

                return resampled

            # Function to limit data points to a maximum number (e.g., 50)
            def limit_data_points(df, max_points=50):
                if len(df) <= max_points:
                    return df

                # Calculate step size to get approximately max_points
                step = len(df) // max_points

                # Return evenly spaced rows
                return df.iloc[::step, :]

            # 1. First visualization: Mill throughput chart (from production_lines PG table)
            try:
                @st.cache_data(ttl=300)
                def load_mill_throughput_pg(_db):
                    engine = create_engine(_db)
                    mill_cols = [f'n{m}num线自磨机处理量t_h' for m in range(1, 7)]
                    quoted = ', '.join([f'"{c}"' for c in mill_cols])
                    query = text(f'SELECT "时间", {quoted} FROM production_lines ORDER BY "时间"')
                    with engine.connect() as conn:
                        df = pd.read_sql(query, conn)
                    df = df.rename(columns={
                        '时间': 'time',
                        **{f'n{m}num线自磨机处理量t_h': f'Mill #{m} (t/h)' for m in range(1, 7)}
                    })
                    df['time'] = pd.to_datetime(df['time'])
                    return df

                mill_data = load_mill_throughput_pg(DB_CONNECTION)

                # Aggregate data based on refresh interval
                mill_data = aggregate_data(mill_data, refresh_interval)

                # Limit the number of data points to improve performance
                mill_data = limit_data_points(mill_data, max_points=50)

                # Mill names (columns 1-6 after time)
                mill_names = [f"Mill #{m} (t/h)" for m in range(1, 7)]

                # Create the figure
                fig = go.Figure()

                # Add each mill's throughput curve
                for mill_name in mill_names:
                    if mill_name in mill_data.columns:
                        fig.add_trace(go.Scatter(
                            x=mill_data['time'],
                            y=mill_data[mill_name],
                            mode='lines',
                            name=mill_name,
                            line=dict(shape='spline')
                        ))

                last_updated_throughput = mill_data['time'].max()
                fig.update_layout(
                    title=f"Mill Throughput - Last updated: {last_updated_throughput.strftime('%I:%M:%S %p %d-%m-%Y')}",
                    xaxis_title="Date and Time",
                    yaxis_title="Throughput Rate - t/h",
                    template="plotly_dark"
                )

                st.plotly_chart(fig, use_container_width=True)

            except Exception as e:
                st.error(f"Error loading mill throughput data: {e}")
                st.warning("No data available. Please check database connection.")

            # 2. Mill mineral data visualization - moved above stockpile visualization
            # 添加缓存装饰器用于数据加载
            @st.cache_data(ttl=600)  # 缓存1小时
            def load_mill_data_last_row(excel_file_path):
                """只加载每个Mill工作表的最后一行数据，大幅提高读取速度"""
                all_mill_data = {}

                # 加载工作簿但不读取数据(data_only=True转换公式为值)
                wb = load_workbook(filename=excel_file_path, read_only=True, data_only=True)

                for i in range(1, 7):
                    sheet_name = f"Mill_{i}"
                    try:
                        if sheet_name in wb.sheetnames:
                            ws = wb[sheet_name]

                            # 获取列名
                            headers = [cell.value for cell in next(ws.rows)]

                            # 直接访问最后一行
                            last_row = None
                            for row in ws.rows:
                                last_row = row  # 持续更新直到最后一行

                            if last_row:
                                # 创建最后一行的字典
                                row_data = {headers[idx]: cell.value for idx, cell in enumerate(last_row) if
                                            idx < len(headers)}

                                # 创建只有一行的DataFrame
                                all_mill_data[sheet_name] = pd.DataFrame([row_data])
                            else:
                                st.warning(f"{sheet_name} 工作表中没有数据")
                        else:
                            st.warning(f"Excel文件中不存在 {sheet_name} 工作表")
                    except Exception as e:
                        st.warning(f"读取 {sheet_name} 数据时出错: {e}")

                wb.close()
                return all_mill_data

            # 缓存图表生成函数
            @st.cache_data(ttl=600)
            def generate_mill_composition_chart(mill_compositions, display_option):
                """生成mill composition图表，使用缓存避免重复生成"""
                # Get all unique composition types across all mills
                all_components = set()
                for compositions in mill_compositions.values():
                    all_components.update(compositions.keys())

                all_components = sorted(list(all_components))

                # Prepare data for each mill
                mills = list(mill_compositions.keys())

                # Create consistent colors for components
                np.random.seed(42)  # For reproducible colors
                component_colors = {}
                colorscale = [
                    "#1f77b4", "#ff7f0e", "#2ca02c", "#d62728", "#9467bd",
                    "#8c564b", "#e377c2", "#7f7f7f", "#bcbd22", "#17becf"
                ]

                for i, component in enumerate(all_components):
                    component_colors[component] = colorscale[i % len(colorscale)]

                # Create bar chart
                fig_mill_ore = go.Figure()

                # Create x-axis labels
                x_labels = mills

                # Create hover texts
                hover_texts = []
                for mill in mills:
                    compositions = mill_compositions.get(mill, {})

                    hover_text = f"<b>{mill}</b><br>"
                    hover_text += "<br>".join(
                        [
                            f"{comp}: {round(compositions.get(comp, 0), 2)} {'tons' if display_option == 'Tonnage' else '%'}"
                            for comp in all_components if comp in compositions]
                    )
                    hover_texts.append(hover_text)

                # Add traces for each component
                for component in all_components:
                    # Get values for this component across mills
                    y_values = []
                    for mill in mills:
                        compositions = mill_compositions.get(mill, {})
                        y_values.append(compositions.get(component, 0))

                    fig_mill_ore.add_trace(
                        go.Bar(
                            x=x_labels,
                            y=y_values,
                            name=component,
                            marker_color=component_colors.get(component, "#CCCCCC"),
                            text=[component for v in y_values],  # Show all component labels
                            textposition="inside"
                        )
                    )

                # Add invisible traces for better hover experience
                for i, mill in enumerate(mills):
                    fig_mill_ore.add_trace(
                        go.Scatter(
                            x=[mill],
                            y=[0],
                            mode="markers",
                            marker=dict(size=0, opacity=0),
                            hoverinfo="text",
                            hovertext=hover_texts[i],
                            showlegend=False
                        )
                    )

                # Update layout
                value_type = "Tons" if display_option == "Tonnage" else "Percentage (%)"
                fig_mill_ore.update_layout(
                    title=f"Mill Composition - {value_type} - Last updated: {datetime.datetime.now().strftime('%I:%M:%S %p %d-%m-%Y')}",
                    xaxis_title="Mills",
                    yaxis_title=f"Composition ({value_type})",
                    barmode='stack',
                    template="plotly_dark",
                    legend_title="Components",
                    hovermode="closest",
                    bargap=0.2,
                    # Adjust these settings to fix the legend position
                    margin=dict(r=150),  # Add right margin to make room for legend
                    legend=dict(
                        orientation="v",  # Vertical legend
                        yanchor="top",  # Anchor to top
                        y=1,  # Position at top
                        xanchor="left",  # Change from "right" to "left"
                        x=1.05,  # Position slightly to the right of the chart (was 1.15)
                    )
                )

                # If showing percentage, set y-axis range to 0-100
                if display_option == "Percentage":
                    fig_mill_ore.update_layout(yaxis=dict(range=[0, 100]))

                return fig_mill_ore

            try:
                # 从 mill_feed 表查询每台磨机的最新组成数据
                @st.cache_data(ttl=300)
                def load_mill_composition_pg(_db):
                    engine = create_engine(_db)
                    query = text("""
                        SELECT DISTINCT ON (mill_num)
                            mill_num, time, mill_composition_tons, mill_composition_pct,
                            mill_composition_properties
                        FROM mill_feed
                        ORDER BY mill_num, time DESC
                    """)
                    with engine.connect() as conn:
                        df = pd.read_sql(query, conn)
                    return df

                # 使用selectbox替代radio按钮，默认选择Tonnage
                display_option = st.selectbox(
                    "Select display type for mill composition:",
                    ["Tonnage", "Percentage"],
                    index=0
                )

                # Determine which column to use based on user selection
                column_to_use = "mill_composition_tons" if display_option == "Tonnage" else "mill_composition_pct"

                all_mill_df = load_mill_composition_pg(DB_CONNECTION)

                def parse_jb(val):
                    if isinstance(val, dict):
                        return val
                    if isinstance(val, str) and val not in ('{}', 'null', ''):
                        try:
                            return json.loads(val)
                        except Exception:
                            return {}
                    return {}

                last_updated_comp = pd.to_datetime(all_mill_df['time']).max() if 'time' in all_mill_df.columns and not all_mill_df.empty else None

                mill_compositions = {}
                mill_properties_by_mill = {}
                for _, row in all_mill_df.iterrows():
                    i = int(row['mill_num'])
                    comp_data = parse_jb(row.get(column_to_use))
                    if comp_data:
                        mill_compositions[f"Mill #{i}"] = comp_data
                    prop_data = parse_jb(row.get('mill_composition_properties'))
                    if prop_data:
                        mill_properties_by_mill[f"Mill #{i}"] = prop_data

                # Mill composition chart
                if mill_compositions:
                    fig_mill_ore = generate_mill_composition_chart(mill_compositions, display_option)
                    if last_updated_comp is not None:
                        value_type = "Tons" if display_option == "Tonnage" else "Percentage (%)"
                        fig_mill_ore.update_layout(
                            title=f"Mill Composition - {value_type} - Last updated: {last_updated_comp.strftime('%I:%M:%S %p %d-%m-%Y')}"
                        )
                    st.plotly_chart(fig_mill_ore, use_container_width=True)
                else:
                    st.warning("No mill composition data available in mill_feed table")

                # ── Weighted Ore Properties section ──────────────────────────
                if mill_properties_by_mill:
                    st.subheader("Mill Feed Weighted Ore Properties", divider="rainbow")

                    prop_options = ["MagFe%", "DTR%", "Fe_Head%", "MagFe_DTR%", "SiO2_DTR%", "Fe_Concentrate%", "P80_IMT_um"]
                    available_props = [p for p in prop_options if any(p in v for v in mill_properties_by_mill.values())]

                    if available_props:
                        # Row 1: Select Mills (full width)
                        # Always offer all 6 mills; default to those with data in current snapshot
                        all_mill_options = [f"Mill #{i}" for i in range(1, 7)]
                        selected_mills_prop = st.multiselect(
                            "Select Mills", all_mill_options, default=all_mill_options, key="mill_prop_select"
                        )

                        # Row 2: Time controls (above property selector)
                        time_range_options = {
                            "Last 1 hour (10-min interval)":   (1,   10),
                            "Last 3 hours (20-min interval)":  (3,   20),
                            "Last 24 hours (1-hour interval)": (24,  60),
                            "Last 3 days (6-hour interval)":   (72, 360),
                        }
                        anchor_default = last_updated_comp.to_pydatetime() if last_updated_comp is not None else datetime.datetime.now()
                        col_range, col_date, col_time = st.columns([2, 1, 1])
                        with col_range:
                            selected_range = st.selectbox(
                                "Trend time range", list(time_range_options.keys()),
                                index=1, key="mill_trend_range"
                            )
                        with col_date:
                            anchor_date = st.date_input(
                                "Anchor date", value=anchor_default.date(), key="trend_anchor_date"
                            )
                        with col_time:
                            anchor_time_val = st.time_input(
                                "Anchor time", value=anchor_default.time(), key="trend_anchor_time"
                            )
                        hours_back, resample_min = time_range_options[selected_range]
                        anchor_dt = datetime.datetime.combine(anchor_date, anchor_time_val)
                        start_dt = anchor_dt - datetime.timedelta(hours=hours_back)

                        # Row 3: Select Property (horizontal radio)
                        selected_prop = st.radio(
                            "Select Property", available_props, horizontal=True, key="mill_prop_radio"
                        )

                        if selected_mills_prop:
                            bar_colors = ["#4C8BF5", "#E8694A", "#56B37F", "#F5C842", "#9B6CF5", "#F5874C"]

                            # ── Chart 1: Single-point bar at anchor time ──────
                            @st.cache_data(ttl=300)
                            def load_mill_props_at_anchor(_db, anchor_dt):
                                engine = create_engine(_db)
                                query = text("""
                                    SELECT DISTINCT ON (mill_num)
                                        mill_num, time, mill_composition_properties
                                    FROM mill_feed
                                    WHERE time <= :anchor
                                    ORDER BY mill_num, time DESC
                                """)
                                with engine.connect() as conn:
                                    df = pd.read_sql(query, conn, params={'anchor': anchor_dt})
                                if df.empty:
                                    return {}
                                result = {}
                                for _, row in df.iterrows():
                                    val = row['mill_composition_properties']
                                    props = val if isinstance(val, dict) else (json.loads(val) if isinstance(val, str) and val else {})
                                    if props:
                                        result[f"Mill #{int(row['mill_num'])}"] = props
                                return result

                            mill_props_anchor = load_mill_props_at_anchor(DB_CONNECTION, anchor_dt)
                            bar_x = [m for m in selected_mills_prop if selected_prop in mill_props_anchor.get(m, {})]
                            bar_y = [mill_props_anchor[m][selected_prop] for m in bar_x]
                            ts_label = anchor_dt.strftime('%I:%M:%S %p %d-%m-%Y')

                            @st.cache_data(ttl=300)
                            def load_mill_properties_ts(_db, start_dt, end_dt):
                                engine = create_engine(_db)
                                query = text("""
                                    SELECT time, mill_num, mill_composition_properties
                                    FROM mill_feed
                                    WHERE time >= :start AND time <= :end
                                    ORDER BY time, mill_num
                                """)
                                with engine.connect() as conn:
                                    df = pd.read_sql(query, conn, params={'start': start_dt, 'end': end_dt})
                                if df.empty:
                                    return df
                                df['time'] = pd.to_datetime(df['time'])
                                df['mill_composition_properties'] = df['mill_composition_properties'].apply(
                                    lambda x: x if isinstance(x, dict) else (json.loads(x) if isinstance(x, str) and x else {})
                                )
                                return df

                            ts_df = load_mill_properties_ts(DB_CONNECTION, start_dt, anchor_dt)

                            # Pre-process trend traces before creating the subplot so we know
                            # whether data exists and can size the figure accordingly
                            trend_traces = []
                            if not ts_df.empty:
                                for idx, mill_name in enumerate(selected_mills_prop):
                                    mill_num_val = int(mill_name.split("#")[1])
                                    m_df = ts_df[ts_df['mill_num'] == mill_num_val].copy()
                                    if m_df.empty:
                                        continue
                                    m_df['prop_val'] = m_df['mill_composition_properties'].apply(
                                        lambda x: x.get(selected_prop) if isinstance(x, dict) else None
                                    )
                                    m_df = m_df.dropna(subset=['prop_val'])
                                    if m_df.empty:
                                        continue
                                    m_df = (m_df.set_index('time')
                                              .resample(f'{resample_min}min')['prop_val']
                                              .mean().reset_index())
                                    if not m_df.empty:
                                        trend_traces.append((mill_name, idx, m_df))

                            trend_has_data = bool(trend_traces)

                            # Row heights and figure height depend on data availability
                            if trend_has_data:
                                row_h = [0.28, 0.72]
                                fig_height = 560
                            else:
                                row_h = [0.65, 0.35]
                                fig_height = 320

                            # ── Combined subplot: bar (top) + line/message (bottom) ───
                            fig_combined = make_subplots(
                                rows=2, cols=1,
                                row_heights=row_h,
                                vertical_spacing=0.06,
                                shared_xaxes=False,
                            )

                            # Row 1: bar chart
                            for idx, (mill, val) in enumerate(zip(bar_x, bar_y)):
                                fig_combined.add_trace(go.Bar(
                                    x=[mill], y=[val],
                                    name=mill,
                                    marker_color=bar_colors[idx % len(bar_colors)],
                                    text=[f"<b>{val:.2f}</b>"],
                                    textposition="outside",
                                    width=0.4,
                                    showlegend=False,
                                ), row=1, col=1)

                            # Row 2: trend lines or no-data message
                            if trend_has_data:
                                for mill_name, idx, m_df in trend_traces:
                                    fig_combined.add_trace(go.Scatter(
                                        x=m_df['time'],
                                        y=m_df['prop_val'],
                                        mode='lines+markers',
                                        name=mill_name,
                                        line=dict(color=bar_colors[idx % len(bar_colors)], width=2),
                                        marker=dict(size=6),
                                    ), row=2, col=1)
                            else:
                                row2_center = row_h[1] * (1 - 0.06) / 2
                                fig_combined.add_annotation(
                                    text=(f"No trend data for "
                                          f"{start_dt.strftime('%d-%m-%Y %H:%M')} → "
                                          f"{anchor_dt.strftime('%d-%m-%Y %H:%M')}"),
                                    xref="paper", yref="paper",
                                    x=0.5, y=row2_center,
                                    showarrow=False,
                                    font=dict(size=13, color="#888888"),
                                    xanchor='center', yanchor='middle',
                                )
                                fig_combined.update_yaxes(visible=False, row=2, col=1)
                                fig_combined.update_xaxes(visible=False, row=2, col=1)

                            # Subtitle y position: just below main title regardless of fig height
                            plot_h_px = fig_height - 85 - 50
                            subtitle_top_px = 0.03 * fig_height + 17 + 8
                            subtitle_y = (fig_height - subtitle_top_px - 50) / plot_h_px

                            fig_combined.update_layout(
                                title=dict(
                                    text=f"Mill Feed Weighted - {selected_prop}",
                                    font=dict(size=17),
                                    y=0.97,
                                    yanchor='top',
                                ),
                                template="plotly_dark",
                                height=fig_height,
                                hovermode="x unified",
                                legend_title="Mills",
                                bargap=0.35,
                                margin=dict(t=85, b=50, l=60, r=20),
                            )
                            fig_combined.add_annotation(
                                text=f"<b>Bars: snapshot at {ts_label}  ·  Trend: {selected_range}</b>",
                                xref="paper", yref="paper",
                                x=0, y=subtitle_y,
                                showarrow=False,
                                font=dict(size=14),
                                xanchor='left', yanchor='top',
                            )
                            fig_combined.update_yaxes(title_text=selected_prop, row=1, col=1,
                                                      range=[0, max(bar_y) * 1.3] if bar_y else [0, 1])
                            fig_combined.update_xaxes(title_text="Mills", row=1, col=1)
                            if trend_has_data:
                                fig_combined.update_yaxes(title_text=selected_prop, row=2, col=1)
                                fig_combined.update_xaxes(title_text="Time", row=2, col=1)

                            st.plotly_chart(fig_combined, use_container_width=True)

            except Exception as e:
                st.error(f"Error loading mill composition data: {e}")
                st.warning("No data available. Please check database connection.")

            # 3. Stockpile data visualization - Moved below mill composition
            def generate_stockpile_data(dates):
                np.random.seed(42)
                stockpiles = ["S1", "S2", "S3", "S4", "S5", "S6"]
                data_frames = []

                for stockpile in stockpiles:
                    data = {
                        "Date": dates,
                        "Stockpile": [stockpile] * len(dates),
                        "MagFe%": np.random.uniform(20, 30, len(dates)),
                        "DTR%": np.random.uniform(25, 35, len(dates)),
                        "MagFe_DTR%": np.random.uniform(15, 25, len(dates)),
                        "Fe_DTR%": np.random.uniform(60, 70, len(dates)),
                        "Fe_Head%": np.random.uniform(30, 40, len(dates)),
                        "Fe_Tail%": np.random.uniform(5, 15, len(dates)),
                        "SiO2_DTR%": np.random.uniform(3, 8, len(dates)),
                        "CaO_Head%": np.random.uniform(1, 5, len(dates)),
                        "MT_Index": np.random.uniform(0.5, 1.5, len(dates)),
                        "Oxidation": np.random.uniform(0.1, 0.9, len(dates)),
                        "Strata": np.random.uniform(1, 5, len(dates)),
                        "PEN_RATE": np.random.uniform(10, 20, len(dates)),
                        "P80_IMT_um": np.random.uniform(300, 800, len(dates))
                    }
                    df = pd.DataFrame(data)
                    data_frames.append(df)

                return data_frames

            # Generate example dates for stockpile data
            dates_mill = pd.date_range(start="2024-10-06", end="2024-12-29", freq="7D")
            stockpile_dfs = generate_stockpile_data(dates_mill)

            selected_stockpiles = st.multiselect('Select Stockpiles to plot',
                                                 ["S1", "S2", "S3", "S4", "S5", "S6"],
                                                 default=["S1", "S2", "S3"])

            # User selects the type of data to be drawn
            data_types = ["MagFe%", "DTR%", "MagFe_DTR%", "Fe_DTR%", "Fe_Head%",
                          "Fe_Tail%", "SiO2_DTR%", "CaO_Head%", "MT_Index", "Oxidation",
                          "Strata", "PEN_RATE", "P80_IMT_um"]
            selected_data_type = st.radio("Select Data Type", data_types, index=0, horizontal=True)

            # Draw a chart
            fig_mill1 = go.Figure()

            for stockpile in selected_stockpiles:
                stockpile_df = next(df for df in stockpile_dfs if df["Stockpile"].iloc[0] == stockpile)

                fig_mill1.add_trace(go.Scatter(
                    x=stockpile_df["Date"],
                    y=stockpile_df[selected_data_type],
                    mode='lines',
                    name=f'Feed Belt - {stockpile}',
                    line=dict(shape='spline')
                ))

            fig_mill1.update_layout(
                title=f"Stockpile Data - {selected_data_type} - Last updated: {datetime.datetime.now().strftime('%I:%M:%S %p %d-%m-%Y')}",
                xaxis_title="Date",
                yaxis_title=f"{selected_data_type}",
                legend_title="Stockpile and Data Type",
                template="plotly_dark"
            )

            st.plotly_chart(fig_mill1, use_container_width=True)

            # 4. Mill throughput prediction (kept as is from original code)
            selected_mill = st.multiselect('Please Select AG Mill For Throughput Prediction',
                                           ["Mill #1", "Mill #2", "Mill #3", "Mill #4", "Mill #5", "Mill #6"],
                                           default=["Mill #1", "Mill #2", "Mill #3"])
            st.button("Confirm", key="confirm_mill_prediction")

            try:
                # Read the XLSX file
                file_path = './resources/Generated_Mill_Data-test.xlsx'
                mill_prediction_data = pd.read_excel(file_path)

                # Limit the number of data points to improve performance
                mill_prediction_data = limit_data_points(mill_prediction_data, max_points=50)

                # Prepare the plotly figures
                fig_mill_predict = go.Figure()

                # Iterate through selected mills and add traces for each
                for mill in selected_mill:
                    real_column = f'{mill}_Real-Data'
                    xgb_column = f'{mill}_XGB'

                    # Add Real-Data trace
                    fig_mill_predict.add_trace(go.Scatter(
                        x=mill_prediction_data['Time'],
                        y=mill_prediction_data[real_column],
                        mode='lines',
                        name=f'{mill} Real-Data'
                    ))

                    # Add XGB trace
                    fig_mill_predict.add_trace(go.Scatter(
                        x=mill_prediction_data['Time'],
                        y=mill_prediction_data[xgb_column],
                        mode='lines',
                        name=f'{mill} XGB Prediction'
                    ))

                # Customize the layout
                fig_mill_predict.update_layout(
                    title=f"Throughput Predictions (Limited to 50 points) - Last updated: {datetime.datetime.now().strftime('%I:%M:%S %p %d-%m-%Y')}",
                    xaxis_title="Time",
                    yaxis_title="Throughput",
                    legend_title="Mill and Data Type",
                    template="plotly_dark"
                )

                # Show the plot in Streamlit
                st.plotly_chart(fig_mill_predict, use_container_width=True)

            except Exception as e:
                st.error(f"Error loading mill prediction data: {e}")
                st.info("Unable to display mill prediction data")

        elif app_chosen == "AG Mill Performance Forecast":
            ### App 5 ##################
            st.subheader("AG Mill Performance Forecast", divider="rainbow")


        elif app_chosen == "Configure Process Parameters":
            ### App 6 ##################
            st.subheader("Configure Process Parameters", divider="rainbow")
            st.markdown("**Current Ore Processing Flowsheet**")
            image_path   = "./resources/processing flowsheet.png"
            config_path  = "./resources/delay.config"
            changelog_path = "./resources/delay.config.changelog"

            # ── 默认配置（46个参数，与 delay.config 结构一致）─────────────────
            _DEFAULT_CFG = {
                "node0_crusher": {
                    "t1a": 0.5, "t1b": 0.5, "t1c": 0.5, "t1d": 0.5,
                    "t2a": 0.5, "t2b": 0.5, "t2c": 0.5, "t2d": 0.5,
                    "t3a": 0.5, "t3b": 0.5, "t3c": 0.5, "t3d": 0.5,
                    "t4a": 0.5, "t4b": 0.5, "t4c": 0.5, "t4d": 0.5,
                },
                "node1_cvr": {
                    "t5a": 3.0, "t5b": 1.0, "t5c": 3.0, "t5d": 2.0,
                    "t6a": 0.0, "t6b": 0.0,
                },
                "node4_silo": {f"t7{chr(ord('a')+i)}": 10.0 for i in range(18)},
                "node45_mill": {f"t8t9t10_{chr(ord('a')+i)}": 15.0 for i in range(6)},
            }

            def read_config():
                try:
                    with open(config_path, 'r') as f:
                        cfg = json.load(f)
                    # 补全缺失 key（兼容旧版 config）
                    for section, defaults in _DEFAULT_CFG.items():
                        cfg.setdefault(section, {})
                        for k, v in defaults.items():
                            cfg[section].setdefault(k, v)
                    return cfg
                except (FileNotFoundError, json.JSONDecodeError):
                    save_config(_DEFAULT_CFG)
                    return {s: dict(d) for s, d in _DEFAULT_CFG.items()}

            def save_config(config_data):
                os.makedirs(os.path.dirname(config_path), exist_ok=True)
                temp_file = config_path + ".tmp"
                try:
                    with open(temp_file, 'w') as f:
                        json.dump(config_data, f, indent=4)
                    if os.path.exists(config_path):
                        os.remove(config_path)
                    os.rename(temp_file, config_path)
                    return True
                except Exception as e:
                    st.error(f"Error saving configuration: {str(e)}")
                    if os.path.exists(temp_file):
                        os.remove(temp_file)
                    return False

            def append_changelog(old_cfg, new_cfg):
                """追加修改记录到 changelog 文件"""
                import datetime as _dt
                lines = [f"\n[{_dt.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}]"]
                changed = False
                for section in new_cfg:
                    for k, v in new_cfg[section].items():
                        old_v = old_cfg.get(section, {}).get(k)
                        if old_v != v:
                            lines.append(f"  {section}.{k}: {old_v} → {v}")
                            changed = True
                if changed:
                    try:
                        with open(changelog_path, 'a', encoding='utf-8') as f:
                            f.write('\n'.join(lines))
                    except Exception:
                        pass

            # ── Session state 初始化 ──────────────────────────────────────────
            if "original_values" not in st.session_state:
                st.session_state.original_values = read_config()
            if "current_values" not in st.session_state:
                st.session_state.current_values = {s: dict(d) for s, d in st.session_state.original_values.items()}
            if "show_confirm_save" not in st.session_state:
                st.session_state.show_confirm_save = False
            if "button_confirm_delay_time" not in st.session_state:
                st.session_state.button_confirm_delay_time = False

            # ── 流程图显示 ────────────────────────────────────────────────────
            if "annotated_image" not in st.session_state:
                display_image = Image.open(image_path)
            else:
                display_image = st.session_state.annotated_image
            image_zoom(display_image, mode="scroll", size=(700, 600),
                       keep_aspect_ratio=False, keep_resolution=True,
                       zoom_factor=10.0, increment=0.2)

            st.markdown("**Please set the ore transport delay time (minutes)**")

            cv = st.session_state.current_values  # 当前值快捷引用
            new_values = {s: dict(d) for s, d in cv.items()}  # 深拷贝，收集修改后的值

            # 统一步长/格式
            _STEP   = 0.5
            _FMT    = "%.1f"

            # ── Node 0 — Crusher (T1~T4, 4台破碎机 A/B/C/D) ─────────────────
            with st.expander("**Node 0 — Crusher  (T1+T2+T3+T4, 4 lines)**", expanded=False):
                st.caption("T1: Dump Truck→Feed Chute | T2: Gyratory Crusher | T3: Crusher→Bottom Bin | T4: Bottom Bin→Belt Feeder")
                for t_stage, label in [("t1","T1"),("t2","T2"),("t3","T3"),("t4","T4")]:
                    cols = st.columns(4)
                    for idx, suffix in enumerate(['a','b','c','d']):
                        k = f"{t_stage}{suffix}"
                        new_values["node0_crusher"][k] = cols[idx].number_input(
                            f"{label}-{suffix.upper()}  (A/B/C/D)",
                            value=float(cv["node0_crusher"][k]),
                            step=_STEP, format=_FMT,
                            key=f"cfg_{k}"
                        )

            # ── Node 1 — CVR & Long Belt (T5~T6, 共6个参数，每行3列) ──────────
            with st.expander("**Node 1 — CVR & Long Belt  (T5: 4 CVR belts | T6: 2 long belts)**", expanded=False):
                st.caption("T5: Belt Feeder→Long Belt (per CVR) | T6: Long Belt→Stacker (shared per line)")
                node1_params = [
                    ("t5a", "T5 CVR111"),
                    ("t5b", "T5 CVR112"),
                    ("t5c", "T5 CVR113"),
                    ("t5d", "T5 CVR114"),
                    ("t6a", "T6A (CVR111/112)"),
                    ("t6b", "T6B (CVR113/114)"),
                ]
                for row_start in range(0, 6, 3):
                    cols = st.columns(3)
                    for col_w, (k, label) in zip(cols, node1_params[row_start:row_start+3]):
                        new_values["node1_cvr"][k] = col_w.number_input(
                            label, value=float(cv["node1_cvr"][k]),
                            step=_STEP, format=_FMT, key=f"cfg_{k}"
                        )

            # ── Node 4 — Silo (T7, 18个料仓) ─────────────────────────────────
            with st.expander("**Node 4 — Silo  (T7: Tripper → each Silo, 18 silos)**", expanded=False):
                st.caption("Silo 1-9 → Tripper 1 | Silo 10-18 → Tripper 2")
                # 每行3个，共6行，列宽充足保证 +/- 按钮正常显示
                for row_start in range(1, 19, 3):
                    cols = st.columns(3)
                    for col_w, silo_num in zip(cols, range(row_start, min(row_start + 3, 19))):
                        k = 't7' + chr(ord('a') + silo_num - 1)
                        new_values["node4_silo"][k] = col_w.number_input(
                            f"Silo {silo_num}",
                            value=float(cv["node4_silo"][k]),
                            step=_STEP, format=_FMT,
                            key=f"cfg_{k}"
                        )

            # ── Node 4/5 — Mill (T8+T9+T10, 6条磨机线路) ─────────────────────
            with st.expander("**Node 4/5 — Mill  (T8+T9+T10 combined, 6 mill lines)**", expanded=False):
                st.caption("T8: Stockpile→Feeder | T9: Feeder→Mill Feed Belt | T10: Mill Feed Belt→AG Mill")
                # 每行3个，共2行
                mill_items = list(enumerate(['a','b','c','d','e','f']))
                for row_start in range(0, 6, 3):
                    cols = st.columns(3)
                    for col_w, (idx, suffix) in zip(cols, mill_items[row_start:row_start+3]):
                        k = f"t8t9t10_{suffix}"
                        new_values["node45_mill"][k] = col_w.number_input(
                            f"Mill {idx+1}",
                            value=float(cv["node45_mill"][k]),
                            step=_STEP, format=_FMT,
                            key=f"cfg_{k}"
                    )

            # ── 检测是否有变更 ────────────────────────────────────────────────
            values_changed = new_values != st.session_state.original_values

            col1, col2 = st.columns(2)

            # Apply to Diagram — 用各节点代表值（A线）更新流程图标注
            with col1:
                if st.button("Apply to Diagram", type="primary", key="apply_button", use_container_width=True):
                    st.session_state.button_confirm_delay_time = True
                    st.session_state.current_values = {s: dict(d) for s, d in new_values.items()}
                    c0 = new_values["node0_crusher"]
                    c1n = new_values["node1_cvr"]
                    c4 = new_values["node4_silo"]
                    c45 = new_values["node45_mill"]
                    st.session_state.annotated_image = annotate_image(
                        image_path,
                        c0["t1a"], c0["t2a"], c0["t3a"], c0["t4a"],
                        c1n["t5a"], c1n["t6a"],
                        c4["t7a"],
                        c45["t8t9t10_a"], 0.0, 0.0   # T8=combined, T9/T10 placeholder
                    )
                    st.rerun()

            # Save Configuration — 写文件 + changelog
            with col2:
                save_button = st.button("Save Configuration", type="primary", key="save_button",
                                        use_container_width=True, disabled=not values_changed)
                if save_button and values_changed:
                    st.session_state.show_confirm_save = True

            if st.session_state.show_confirm_save:
                st.warning("Are you sure you want to save these changes to the configuration file?")
                col3, col4 = st.columns(2)
                with col3:
                    if st.button("Confirm Save", key="confirm_save", use_container_width=True):
                        if save_config(new_values):
                            append_changelog(st.session_state.original_values, new_values)
                            st.success("Configuration saved successfully!")
                            st.session_state.original_values = {s: dict(d) for s, d in new_values.items()}
                            st.session_state.current_values  = {s: dict(d) for s, d in new_values.items()}
                        st.session_state.show_confirm_save = False
                        st.rerun()
                with col4:
                    if st.button("Cancel", key="cancel_save", use_container_width=True):
                        st.session_state.show_confirm_save = False
                        st.rerun()

        else:  # "Configure Database Connection"
            ### App 7 ##################
            st.subheader("Configure Database Connection", divider="rainbow")
            st.markdown("**Please configure the API json call for the database**")
            ##################################################
            # Define initial JSON data
            initial_data = {
                "database": {
                    "type": "postgresql",
                    "host": "localhost",
                    "port": 5432,
                    "database_name": "my_database",
                    "username": "my_user",
                    "password": "my_password"
                }
            }

            # Dropdown for database type selection
            db_type = st.selectbox("Choose Database Type", ["postgresql", "mysql", "sqlite"], index=0)
            initial_data['database']['type'] = db_type  # Update JSON data with selected type

            # Convert to JSON string for editing
            json_str = json.dumps(initial_data, indent=4)
            json_input = st.text_area("Edit JSON Data", value=json_str, height=300)

            # Confirmation button
            if st.button("Confirm and Update JSON"):
                try:
                    updated_data = json.loads(json_input)
                    st.success("JSON data updated successfully!")
                    st.json(updated_data)
                except json.JSONDecodeError:
                    st.error("Invalid JSON format. Please check your input.")

        st.sidebar.markdown("###")
        st.sidebar.markdown(f"***Welcome {name}!***")
        authenticator.logout("Logout", 'sidebar')


    elif authentication_status is False:
        st.error('Username/password is incorrect')
    elif authentication_status is None:
        st.warning('Please enter your username and password')


    #############################################################################################################################################################################
    st.markdown(
        f"""
            <style>
                .reportview-container .main .block-container{{
                    max-width: 1500px;
                    padding-top: 1rem;
                    padding-right: 1rem;
                    padding-left: 1rem;
                    padding-bottom: 1rem;
                }}

            </style>
            """,
        unsafe_allow_html=True,
    )

    footer = """  
            <style>
                .footer {
                position: fixed;
                left: 0;
                bottom: 0;
                width: 100%;
                background-color: #50575b;
                color: white;
                text-align: center;
                }
            </style>

            <div class="footer">
                <p>Visit us @ https://citicpacificmining.com | © 2024 Copyright Citic Pacific Mining  </p>
            </div>
        """

    st.markdown(footer, unsafe_allow_html=True)

if __name__ == "__main__":
    main()



